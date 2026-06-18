"""
Summary Table Detector — Hybrid table detection using formatting + heuristics.

Primary: Uses cell formatting (bold, borders, indentation) to identify headers,
totals, and section boundaries — the same signals a human uses.

Secondary: Falls back to blank-row/blank-column heuristics for edge cases.

Optional: Uses eparse library for automated table crawling if available.
"""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import formatting_extractor


# ─────────────────────────────────────────────────────────────────
# Row classification using formatting
# ─────────────────────────────────────────────────────────────────

def is_blank_row(ws, row_idx, col_start=1, col_end=None):
    """Check if a row is completely blank within a column range."""
    if col_end is None:
        col_end = ws.max_column or 1
    for c in range(col_start, col_end + 1):
        if ws.cell(row=row_idx, column=c).value is not None:
            return False
    return True


def get_non_blank_row_blocks(ws):
    """
    Find contiguous ranges of rows that contain at least one non-empty cell.
    Returns a list of tuples (start_row, end_row).
    """
    blocks = []
    in_block = False
    start_row = None
    
    max_row = ws.max_row or 1
    for r in range(1, max_row + 1):
        if not is_blank_row(ws, r):
            if not in_block:
                in_block = True
                start_row = r
        else:
            if in_block:
                blocks.append((start_row, r - 1))
                in_block = False
                
    if in_block:
        blocks.append((start_row, max_row))
        
    return blocks


def detect_horizontal_splits(ws_val, row_start, row_end):
    """
    For a given row range, identify column blocks separated by blank columns.
    Returns a list of tuples (col_start, col_end).
    """
    max_col = ws_val.max_column or 1
    active_cols = []
    
    for c in range(1, max_col + 1):
        col_has_data = False
        for r in range(row_start, row_end + 1):
            if ws_val.cell(row=r, column=c).value is not None:
                col_has_data = True
                break
        if col_has_data:
            active_cols.append(c)
            
    if not active_cols:
        return []
        
    blocks = []
    start_col = active_cols[0]
    prev_col = active_cols[0]
    
    for c in active_cols[1:]:
        if c > prev_col + 1:
            blocks.append((start_col, prev_col))
            start_col = c
        prev_col = c
    blocks.append((start_col, prev_col))
    
    return blocks


def is_value_cell(cell_val, cell_formula):
    """Determine if a cell represents a value or formula (as opposed to metadata/labels)."""
    if cell_val is None:
        return False
    if str(cell_formula).startswith('='):
        return True
    if isinstance(cell_val, (int, float)):
        return True
    return False


def classify_table_rows(ws_val, ws_form, row_start, row_end, col_start, col_end):
    """
    Classify rows in a table block using FORMATTING as the primary signal,
    with content heuristics as fallback.
    
    Returns dict with: title_rows, header_rows, data_rows, total_rows, check_rows
    """
    title_rows = []
    header_rows = []
    data_rows = []
    total_rows = []
    check_rows = []
    
    # 1. Analyze each row using formatting signals
    row_analyses = {}
    for r in range(row_start, row_end + 1):
        row_analyses[r] = formatting_extractor.analyze_row_formatting(
            ws_val, r, col_start, col_end
        )
    
    # 2. Identify total and check rows first (they have clear signals)
    for r in range(row_start, row_end + 1):
        analysis = row_analyses[r]
        
        if analysis["is_blank"]:
            continue
        
        first_val = analysis.get("first_value", "")
        
        # Check rows
        if analysis.get("is_check") or "check" in first_val:
            check_rows.append(r)
            continue
        
        # Total rows — detected by formatting (bold) + content ("total")
        if analysis["is_total"]:
            total_rows.append(r)
            continue
        
        # Also detect total rows by SUM formula with blank labels
        has_sum_formula = False
        label_cols_blank = True
        for c in range(col_start, min(col_start + 2, col_end)):
            formula = ws_form.cell(row=r, column=c).value
            if formula is not None and any(x in str(formula).upper() for x in ["SUM(", "SUBTOTAL("]):
                has_sum_formula = True
            if ws_val.cell(row=r, column=c).value is not None:
                label_cols_blank = False
        
        if has_sum_formula and label_cols_blank and r > row_start:
            total_rows.append(r)
            continue
        
        # Check if content says "total" even without bold
        if "total" in first_val or "grand total" in first_val:
            total_rows.append(r)
            continue
    
    # 3. Find the first data row
    data_start_row = None
    for r in range(row_start, row_end + 1):
        if r in total_rows or r in check_rows:
            continue
        
        analysis = row_analyses[r]
        if analysis["is_blank"]:
            continue
        
        # FORMATTING-BASED: If this row has numeric data
        if analysis["has_data"]:
            # Also verify it's not a header (bold + no numeric in some cols)
            if not analysis["is_header"]:
                data_start_row = r
                break
        
        # CONTENT-BASED fallback: Check for numeric values after label columns
        val_start_col = min(col_start + 2, col_end)
        for c in range(val_start_col, col_end + 1):
            val = ws_val.cell(row=r, column=c).value
            formula = ws_form.cell(row=r, column=c).value
            if is_value_cell(val, formula):
                data_start_row = r
                break
        if data_start_row:
            break
    
    if data_start_row is None:
        if row_end > row_start:
            data_start_row = row_start + 1
        else:
            data_start_row = row_start
    
    # 4. Classify rows before data_start_row as title or header
    for r in range(row_start, data_start_row):
        if r in total_rows or r in check_rows:
            continue
        
        analysis = row_analyses[r]
        if analysis["is_blank"]:
            continue
        
        # FORMATTING-BASED header detection
        if analysis["is_header"]:
            header_rows.append(r)
        elif analysis["is_section_title"]:
            title_rows.append(r)
        else:
            # CONTENT-BASED fallback
            row_vals = [ws_val.cell(row=r, column=c).value for c in range(col_start, col_end + 1)]
            non_empty = [v for v in row_vals if v is not None]
            
            if len(non_empty) == 1:
                title_rows.append(r)
            elif len(non_empty) > 1:
                # Check if most values are text (header indicator)
                text_count = sum(1 for v in non_empty if isinstance(v, str))
                if text_count / len(non_empty) > 0.5:
                    header_rows.append(r)
                else:
                    header_rows.append(r)
    
    # Ensure at least one header row
    if not header_rows and data_start_row > row_start:
        last_pre_data = data_start_row - 1
        if last_pre_data in title_rows:
            title_rows.remove(last_pre_data)
        header_rows.append(last_pre_data)
    
    # 5. Classify remaining rows as data
    for r in range(data_start_row, row_end + 1):
        if r in total_rows or r in check_rows:
            continue
        analysis = row_analyses[r]
        if not analysis["is_blank"]:
            data_rows.append(r)
    
    return {
        "title_rows": title_rows,
        "header_rows": header_rows,
        "data_rows": data_rows,
        "total_rows": total_rows,
        "check_rows": check_rows,
    }


def build_disambiguated_headers(ws_val, header_rows, col_start, col_end):
    """
    Build column headers. If there are multiple header rows (multi-header),
    combine them to avoid duplicates.
    """
    headers = []
    
    if not header_rows:
        for c in range(col_start, col_end + 1):
            headers.append(f"Column_{get_column_letter(c)}")
        return headers
        
    if len(header_rows) == 1:
        hr = header_rows[0]
        for c in range(col_start, col_end + 1):
            val = ws_val.cell(row=hr, column=c).value
            headers.append(str(val).strip() if val is not None else "")
        return headers
        
    # Multi-row header: combine non-empty parts
    for c in range(col_start, col_end + 1):
        parts = []
        for hr in header_rows:
            val = ws_val.cell(row=hr, column=c).value
            if val is not None:
                parts.append(str(val).strip())
        
        unique_parts = []
        for p in parts:
            if not unique_parts or unique_parts[-1] != p:
                unique_parts.append(p)
                
        header_name = " ".join(unique_parts)
        headers.append(header_name)
        
    return headers


# ─────────────────────────────────────────────────────────────────
# eparse integration (optional)
# ─────────────────────────────────────────────────────────────────

def _try_eparse_detection(file_path, sheet_name):
    """
    Attempt to use eparse library for automatic table detection.
    Returns a list of detected table ranges or None if eparse is not available.
    """
    try:
        from eparse.core import HotSheet
        hs = HotSheet(file_path, sheet_name=sheet_name)
        # eparse returns structured table data
        tables = []
        if hasattr(hs, 'tables'):
            for t in hs.tables:
                tables.append({
                    "headers": t.headers if hasattr(t, 'headers') else [],
                    "row_count": len(t.rows) if hasattr(t, 'rows') else 0,
                })
        return tables if tables else None
    except Exception as e:
        # eparse may not work on all file types or sheet layouts
        return None


# ─────────────────────────────────────────────────────────────────
# Main extraction function
# ─────────────────────────────────────────────────────────────────

def extract_tables_from_sheet(ws_val, ws_form, pivot_tables_meta=None, wb=None):
    """
    Scans the worksheet and extracts all tabular structures.
    
    Uses formatting-based detection as primary, with blank-row/column 
    heuristics as secondary. Integrates pivot metadata for pivot table boundaries.
    """
    detected_tables = []
    
    # Extract sheet-level formatting metadata
    sheet_fmt_meta = formatting_extractor.extract_sheet_metadata(ws_val, wb)
    
    # 1. Mark pivot table ranges
    pivot_ranges = []
    if pivot_tables_meta:
        for pt in pivot_tables_meta:
            pt_range = pt.get("table_range", "")
            if pt_range:
                pivot_ranges.append(pt_range)
                
    # 2. Get all row blocks that contain data
    row_blocks = get_non_blank_row_blocks(ws_val)
    
    # Find start of the first pivot table if present
    min_pt_start = 9999
    if pivot_tables_meta:
        for pt in pivot_tables_meta:
            pt_range = pt.get("table_range", "")
            if pt_range:
                parts = pt_range.split(':')
                pt_start_r = int(''.join(filter(str.isdigit, parts[0])))
                if pt_start_r < min_pt_start:
                    min_pt_start = pt_start_r
                    
    # Filter out title/filter blocks
    table_row_blocks = []
    for start_r, end_r in row_blocks:
        if pivot_tables_meta and end_r < min_pt_start:
            continue
            
        if start_r <= 3 and end_r <= 3:
            cells_count = 0
            for r in range(start_r, end_r + 1):
                for c in range(1, (ws_val.max_column or 1) + 1):
                    if ws_val.cell(row=r, column=c).value is not None:
                        cells_count += 1
            if cells_count < 4:
                continue
        table_row_blocks.append((start_r, end_r))
        
    # 3. Process each row block
    for start_r, end_r in table_row_blocks:
        # Check for pivot table overlap
        is_pivot_block = False
        matching_pivot = None
        for pt in pivot_tables_meta or []:
            pt_range = pt.get("table_range", "")
            if pt_range:
                parts = pt_range.split(':')
                pt_start_r = int(''.join(filter(str.isdigit, parts[0])))
                pt_end_r = int(''.join(filter(str.isdigit, parts[1])))
                if max(start_r, pt_start_r) <= min(end_r, pt_end_r):
                    is_pivot_block = True
                    matching_pivot = pt
                    break
                    
        if is_pivot_block and matching_pivot:
            # Parse pivot table using boundaries from pivot metadata
            pt_range = matching_pivot["table_range"]
            parts = pt_range.split(':')
            col_start_letter = ''.join(filter(str.isalpha, parts[0]))
            col_end_letter = ''.join(filter(str.isalpha, parts[1]))
            col_start = column_index_from_string(col_start_letter)
            col_end = column_index_from_string(col_end_letter)
            pt_start_r = int(''.join(filter(str.isdigit, parts[0])))
            pt_end_r = int(''.join(filter(str.isdigit, parts[1])))
            
            row_classification = classify_table_rows(ws_val, ws_form, pt_start_r, pt_end_r, col_start, col_end)
            
            t_name = matching_pivot["pivot_table_name"]
            
            detected_tables.append({
                "table_name": t_name,
                "table_type": "pivot_table",
                "section_title": "Exhibit 5 Pivot Summary" if "GVUL" in t_name or "Exb" in t_name else t_name,
                "table_range": pt_range,
                "col_start": col_start,
                "col_end": col_end,
                "row_start": pt_start_r,
                "row_end": pt_end_r,
                "row_classification": row_classification,
                "pivot_meta": matching_pivot,
                "formatting_metadata": {
                    "merged_cells": sheet_fmt_meta.get("merged_cells", []),
                },
            })
            continue
            
        # Detect horizontal splits
        col_blocks = detect_horizontal_splits(ws_val, start_r, end_r)
        
        for col_start, col_end in col_blocks:
            # Classify rows using formatting-enhanced detection
            row_classification = classify_table_rows(ws_val, ws_form, start_r, end_r, col_start, col_end)
            
            # Extract section title
            section_title = ""
            title_rows = row_classification["title_rows"]
            if title_rows:
                title_val = ws_val.cell(row=title_rows[0], column=col_start).value
                if title_val is not None:
                    section_title = str(title_val).strip()
                    
            # Build headers
            header_rows = row_classification["header_rows"]
            headers = build_disambiguated_headers(ws_val, header_rows, col_start, col_end)
            
            # Detect hierarchy from indentation
            data_rows = row_classification["data_rows"]
            hierarchy = []
            if data_rows:
                hierarchy = formatting_extractor.detect_row_hierarchy(
                    ws_val, data_rows, col_start
                )
            
            # Table name
            if section_title:
                table_name = section_title
            else:
                table_name = f"Table_{get_column_letter(col_start)}_{start_r}"
                
            table_name = table_name.replace(":", "").replace("\n", " ").strip()
            
            # Table type
            table_type = "section_summary"
            if any(h in ["Gross Reserve", "Net Reserve", "LFG Gross GA STAT Reserve"] for h in headers):
                table_type = "formula_summary"
                
            table_range = f"{get_column_letter(col_start)}{start_r}:{get_column_letter(col_end)}{end_r}"
            
            detected_tables.append({
                "table_name": table_name,
                "table_type": table_type,
                "section_title": section_title,
                "table_range": table_range,
                "col_start": col_start,
                "col_end": col_end,
                "row_start": start_r,
                "row_end": end_r,
                "row_classification": row_classification,
                "headers": headers,
                "hierarchy": hierarchy,
                "formatting_metadata": {
                    "merged_cells": [
                        mc for mc in sheet_fmt_meta.get("merged_cells", [])
                    ],
                },
            })
            
    return detected_tables
