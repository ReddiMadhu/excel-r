"""
Structural Summary — Generate a compact, token-efficient JSON skeleton of a workbook.

This skeleton provides the LLM with full context about workbook layout, headers, 
formulas, and sample data without exceeding token limitations.
"""
import os
import openpyxl
from openpyxl.utils import get_column_letter


def generate_workbook_skeleton(file_path, sheet_types, wb_val, wb_form, raw_column_maps, detected_tables):
    """
    Generate a compact JSON skeleton of the workbook structure.
    
    Parameters:
        file_path (str): Path to the Excel file.
        sheet_types (dict): Mapping of sheet_name -> sheet_type.
        wb_val (openpyxl.Workbook): Workbook loaded with data_only=True.
        wb_form (openpyxl.Workbook): Workbook loaded with data_only=False.
        raw_column_maps (dict): Raw data columns mapping.
        detected_tables (list): List of tables detected across the summary sheets.
        
    Returns:
        dict: The workbook skeleton metadata.
    """
    skeleton = {
        "file_name": os.path.basename(file_path),
        "workbook_metadata": {
            "sheets": [],
            "named_ranges": [],
            "external_links": [],
        }
    }
    
    # 1. Workbook-level metadata (defined names / links)
    # Extract named ranges
    try:
        if wb_val.defined_names:
            for name, dn in list(wb_val.defined_names.items())[:20]:  # Cap at 20
                try:
                    skeleton["workbook_metadata"]["named_ranges"].append({
                        "name": name,
                        "value": dn.value
                    })
                except Exception:
                    pass
    except Exception:
        pass
        
    # Extract external links
    try:
        if hasattr(wb_val, '_external_links') and wb_val._external_links:
            for link in wb_val._external_links:
                try:
                    if hasattr(link, 'file_link') and link.file_link:
                        skeleton["workbook_metadata"]["external_links"].append(str(link.file_link))
                    elif hasattr(link, 'Target'):
                        skeleton["workbook_metadata"]["external_links"].append(str(link.Target))
                except Exception:
                    pass
    except Exception:
        pass

    # 2. Sheet summaries
    for s_name, s_type in sheet_types.items():
        ws_val = wb_val[s_name]
        ws_form = wb_form[s_name]
        
        sheet_info = {
            "sheet_name": s_name,
            "sheet_type": s_type,
            "dimensions": f"A1:{get_column_letter(ws_val.max_column or 1)}{ws_val.max_row or 1}",
            "row_count": ws_val.max_row or 0,
            "column_count": ws_val.max_column or 0,
        }
        
        if s_type == "raw_data":
            # Just describe raw sheets (no detailed cell dumps needed)
            col_headers = []
            if s_name in raw_column_maps:
                col_headers = list(raw_column_maps[s_name].values())
            else:
                # Fallback: read row 1 headers
                for col_idx in range(1, min(ws_val.max_column or 1, 30) + 1):
                    val = ws_val.cell(row=1, column=col_idx).value
                    if val is not None:
                        col_headers.append(str(val).strip())
            sheet_info["raw_columns"] = col_headers[:50]  # Cap at 50 columns to save tokens
            sheet_info["description"] = "Flat database sheet containing raw records."
            
        elif s_type == "summary_report":
            sheet_info["tables"] = []
            
            # Find tables detected on this sheet
            sheet_tables = [t for t in detected_tables if t.get("table_range") and ws_val.cell(row=t["row_start"], column=t["col_start"]).parent.title == s_name]
            
            for tbl in sheet_tables:
                t_name = tbl["table_name"]
                t_type = tbl["table_type"]
                t_range = tbl["table_range"]
                
                row_cls = tbl["row_classification"]
                col_start = tbl["col_start"]
                col_end = tbl["col_end"]
                
                header_rows = row_cls.get("header_rows", [])
                data_rows = row_cls.get("data_rows", [])
                total_rows = row_cls.get("total_rows", [])
                check_rows = row_cls.get("check_rows", [])
                
                headers = tbl.get("headers", [])
                if not headers:
                    # Fallback or pivot table headers
                    if header_rows:
                        headers = [str(ws_val.cell(row=header_rows[-1], column=c).value or f"Col_{get_column_letter(c)}") for c in range(col_start, col_end + 1)]
                    else:
                        headers = [f"Col_{get_column_letter(c)}" for c in range(col_start, col_end + 1)]
                
                # Build columns summary
                columns_summary = []
                for c_idx, col_name in enumerate(headers, col_start):
                    col_letter = get_column_letter(c_idx)
                    
                    # Find first formula in data rows
                    first_formula = None
                    for r in data_rows:
                        val = ws_form.cell(row=r, column=c_idx).value
                        if val is not None and str(val).startswith('='):
                            first_formula = str(val)
                            break
                            
                    # Extract sample values (first 2 rows)
                    sample_vals = []
                    for r in data_rows[:2]:
                        val = ws_val.cell(row=r, column=c_idx).value
                        if val is not None:
                            sample_vals.append(str(val))
                            
                    col_info = {
                        "column_name": col_name,
                        "excel_column": col_letter,
                        "first_formula": first_formula or "",
                        "sample_values": sample_vals
                    }
                    columns_summary.append(col_info)
                
                # Build totals summary (value/formula per column on total rows)
                totals_summary = []
                for r in total_rows:
                    total_row_data = {}
                    label = str(ws_val.cell(row=r, column=col_start).value or "Total")
                    total_row_data["row_index"] = r
                    total_row_data["row_label"] = label
                    total_row_data["column_values"] = {}
                    
                    for c_idx, col_name in enumerate(headers, col_start):
                        val_v = ws_val.cell(row=r, column=c_idx).value
                        val_f = ws_form.cell(row=r, column=c_idx).value
                        
                        val_repr = ""
                        if val_f is not None and str(val_f).startswith('='):
                            val_repr = str(val_f)
                        elif val_v is not None:
                            val_repr = str(val_v)
                        total_row_data["column_values"][col_name] = val_repr
                        
                    totals_summary.append(total_row_data)
                    
                table_info = {
                    "table_name": t_name,
                    "table_type": t_type,
                    "range": t_range,
                    "section_title": tbl.get("section_title", ""),
                    "row_classification_summary": {
                        "header_rows": header_rows,
                        "data_rows_count": len(data_rows),
                        "data_rows_range": f"{data_rows[0]}:{data_rows[-1]}" if data_rows else "",
                        "total_rows": total_rows,
                        "check_rows": check_rows,
                    },
                    "columns": columns_summary,
                    "totals": totals_summary,
                }
                
                # If it's a pivot table, inject its XML metadata
                if t_type == "pivot_table" and "pivot_meta" in tbl:
                    pm = tbl["pivot_meta"]
                    table_info["pivot_metadata"] = {
                        "pivot_table_name": pm.get("pivot_table_name"),
                        "data_source_sheet": pm.get("data_source_sheet"),
                        "row_fields": pm.get("row_fields", []),
                        "column_fields": pm.get("column_fields", []),
                        "value_fields": pm.get("value_fields", []),
                    }
                    
                sheet_info["tables"].append(table_info)
                
        skeleton["workbook_metadata"]["sheets"].append(sheet_info)
        
    return skeleton
