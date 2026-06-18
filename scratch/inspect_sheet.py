import openpyxl

wb_path = r"c:\Users\madhu\Desktop\excelrationlization\input files\data\input\LNBAR 2024 EB Reserve Deatils.xlsx"
wb = openpyxl.load_workbook(wb_path, data_only=False)
print("Sheet names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\nSheet: {sheetname}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Let's inspect rows 1 to 15, cols I to K (columns 9 to 11)
    print("Cells in columns I to K (9 to 11), rows 1 to 15:")
    for r in range(1, 16):
        row_vals = []
        for c in range(9, 12):
            cell = ws.cell(row=r, column=c)
            row_vals.append(f"{cell.coordinate}: {cell.value}")
        print(" | ".join(row_vals))
