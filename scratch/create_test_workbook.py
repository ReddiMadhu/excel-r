"""
Create a test Excel workbook with 3 sheets demonstrating chained formula lineage:
  Sheet A (SQL_data)  →  Sheet B (Summary)  →  Sheet C (FinalReport)

- Sheet A: Raw data with Employee, Department, Region, Salary, Bonus
- Sheet B: Summary table using SUMIFS from Sheet A (salary/bonus totals by department)
- Sheet C: Final report using SUM/formulas from Sheet B (grand totals, ratios, checks)
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# Sheet A: "SQL_data" — raw data source
# ─────────────────────────────────────────────
ws_a = wb.active
ws_a.title = "SQL_data"

headers_a = ["Employee_ID", "Employee_Name", "Department", "Region", "Salary", "Bonus", "Years_of_Service"]
for c, h in enumerate(headers_a, 1):
    ws_a.cell(row=1, column=c, value=h)

raw_data = [
    [1001, "Alice Johnson",   "Engineering",  "North", 95000,  12000, 5],
    [1002, "Bob Smith",       "Engineering",  "South", 88000,  10000, 3],
    [1003, "Carol Davis",     "Engineering",  "North", 102000, 15000, 7],
    [1004, "Dan Wilson",      "Marketing",    "East",  72000,  8000,  4],
    [1005, "Eva Martinez",    "Marketing",    "West",  68000,  7500,  2],
    [1006, "Frank Lee",       "Marketing",    "East",  75000,  9000,  6],
    [1007, "Grace Kim",       "Finance",      "North", 85000,  11000, 4],
    [1008, "Henry Park",      "Finance",      "South", 92000,  13000, 8],
    [1009, "Iris Chen",       "Finance",      "West",  78000,  9500,  3],
    [1010, "Jack Brown",      "HR",           "North", 65000,  6000,  2],
    [1011, "Karen White",     "HR",           "East",  62000,  5500,  1],
    [1012, "Leo Garcia",      "HR",           "South", 67000,  7000,  5],
    [1013, "Mia Patel",       "Engineering",  "East",  98000,  14000, 6],
    [1014, "Noah Taylor",     "Marketing",    "North", 71000,  8500,  3],
    [1015, "Olivia Nguyen",   "Finance",      "East",  89000,  12500, 5],
]

for r_idx, row_data in enumerate(raw_data, 2):
    for c_idx, val in enumerate(row_data, 1):
        ws_a.cell(row=r_idx, column=c_idx, value=val)


# ─────────────────────────────────────────────
# Sheet B: "Summary" — aggregates from Sheet A
# ─────────────────────────────────────────────
ws_b = wb.create_sheet("Summary")

# Title row
ws_b.cell(row=1, column=1, value="Department Compensation Summary")
ws_b.merge_cells("A1:F1")

# Headers (row 2)
headers_b = ["Department", "Total Salary", "Total Bonus", "Total Compensation", "Headcount", "Avg Salary"]
for c, h in enumerate(headers_b, 1):
    ws_b.cell(row=2, column=c, value=h)

# Data rows (rows 3-6): one per department
departments = ["Engineering", "Finance", "HR", "Marketing"]
for i, dept in enumerate(departments):
    r = i + 3
    ws_b.cell(row=r, column=1, value=dept)
    # B: Total Salary = SUMIFS(SQL_data!E:E, SQL_data!C:C, A<r>)
    ws_b.cell(row=r, column=2).value = f"=SUMIFS(SQL_data!E:E,SQL_data!C:C,Summary!$A{r})"
    # C: Total Bonus = SUMIFS(SQL_data!F:F, SQL_data!C:C, A<r>)
    ws_b.cell(row=r, column=3).value = f"=SUMIFS(SQL_data!F:F,SQL_data!C:C,Summary!$A{r})"
    # D: Total Compensation = B + C (local cross-column reference)
    ws_b.cell(row=r, column=4).value = f"=B{r}+C{r}"
    # E: Headcount = COUNTIFS(SQL_data!C:C, A<r>)
    ws_b.cell(row=r, column=5).value = f"=COUNTIFS(SQL_data!C:C,Summary!$A{r})"
    # F: Avg Salary = B / E
    ws_b.cell(row=r, column=6).value = f"=B{r}/E{r}"

# Total row (row 7)
ws_b.cell(row=7, column=1, value="Total")
ws_b.cell(row=7, column=2).value = "=SUM(B3:B6)"
ws_b.cell(row=7, column=3).value = "=SUM(C3:C6)"
ws_b.cell(row=7, column=4).value = "=SUM(D3:D6)"
ws_b.cell(row=7, column=5).value = "=SUM(E3:E6)"
ws_b.cell(row=7, column=6).value = "=B7/E7"

# Check row (row 8): verify total salary matches raw data
ws_b.cell(row=8, column=1, value="Check")
ws_b.cell(row=8, column=2).value = "=SUM(SQL_data!E:E)-B7"
ws_b.cell(row=8, column=3).value = "=SUM(SQL_data!F:F)-C7"
ws_b.cell(row=8, column=4).value = "-"
ws_b.cell(row=8, column=5).value = "-"
ws_b.cell(row=8, column=6).value = "-"

# ── Second table on Summary sheet (columns H-K), rows 2-8 ──
# This is a region-level breakdown (side-by-side horizontal table)
ws_b.cell(row=1, column=8, value="Region Compensation Summary")
ws_b.merge_cells("H1:K1")

headers_b2 = ["Region", "Total Salary", "Total Bonus", "Headcount"]
for c, h in enumerate(headers_b2, 8):
    ws_b.cell(row=2, column=c, value=h)

regions = ["North", "South", "East", "West"]
for i, region in enumerate(regions):
    r = i + 3
    ws_b.cell(row=r, column=8, value=region)
    ws_b.cell(row=r, column=9).value = f"=SUMIFS(SQL_data!E:E,SQL_data!D:D,Summary!$H{r})"
    ws_b.cell(row=r, column=10).value = f"=SUMIFS(SQL_data!F:F,SQL_data!D:D,Summary!$H{r})"
    ws_b.cell(row=r, column=11).value = f"=COUNTIFS(SQL_data!D:D,Summary!$H{r})"

ws_b.cell(row=7, column=8, value="Total")
ws_b.cell(row=7, column=9).value = "=SUM(I3:I6)"
ws_b.cell(row=7, column=10).value = "=SUM(J3:J6)"
ws_b.cell(row=7, column=11).value = "=SUM(K3:K6)"

ws_b.cell(row=8, column=8, value="Check")
ws_b.cell(row=8, column=9).value = "=SUM(SQL_data!E:E)-I7"
ws_b.cell(row=8, column=10).value = "=SUM(SQL_data!F:F)-J7"
ws_b.cell(row=8, column=11).value = "=SUM(SQL_data!K:K)-K7"


# ─────────────────────────────────────────────
# Sheet C: "FinalReport" — references Sheet B (Summary)
# This demonstrates B→C transitive lineage
# ─────────────────────────────────────────────
ws_c = wb.create_sheet("FinalReport")

# Title
ws_c.cell(row=1, column=1, value="Executive Compensation Report")
ws_c.merge_cells("A1:E1")

# Headers (row 2)
headers_c = ["Department", "Total Compensation", "% of Grand Total", "Compensation per Head", "Rating"]
for c, h in enumerate(headers_c, 1):
    ws_c.cell(row=2, column=c, value=h)

# Data rows referencing Summary sheet columns
for i, dept in enumerate(departments):
    r = i + 3
    ws_c.cell(row=r, column=1, value=dept)
    # B: Total Compensation — references Summary!D (which itself is B+C from SQL_data)
    ws_c.cell(row=r, column=2).value = f"=Summary!D{r}"
    # C: % of Grand Total — references Summary!D / Summary!D7
    ws_c.cell(row=r, column=3).value = f"=Summary!D{r}/Summary!$D$7"
    # D: Compensation per Head — references Summary!D / Summary!E
    ws_c.cell(row=r, column=4).value = f"=Summary!D{r}/Summary!E{r}"
    # E: Rating — IF formula based on percentage
    ws_c.cell(row=r, column=5).value = f'=IF(C{r}>0.3,"High",IF(C{r}>0.2,"Medium","Low"))'

# Total row
ws_c.cell(row=7, column=1, value="Total")
ws_c.cell(row=7, column=2).value = "=SUM(B3:B6)"
ws_c.cell(row=7, column=3).value = "=SUM(C3:C6)"
ws_c.cell(row=7, column=4).value = "=B7/SUM(Summary!E3:E6)"
ws_c.cell(row=7, column=5).value = "-"

# Check row — verify grand total matches Summary
ws_c.cell(row=8, column=1, value="Check")
ws_c.cell(row=8, column=2).value = "=Summary!D7-B7"
ws_c.cell(row=8, column=3).value = "=1-C7"
ws_c.cell(row=8, column=4).value = "-"
ws_c.cell(row=8, column=5).value = "-"


# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
out_path = os.path.join("data", "input", "Test_3Sheet_Lineage.xlsx")
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"Created test workbook: {os.path.abspath(out_path)}")
print(f"  Sheet A (SQL_data): {len(raw_data)} data rows, {len(headers_a)} columns — raw source")
print(f"  Sheet B (Summary):  2 tables (Dept A:F, Region H:K) — SUMIFS/COUNTIFS from Sheet A")
print(f"  Sheet C (FinalReport): 1 table — formulas referencing Sheet B (transitive lineage A→B→C)")
