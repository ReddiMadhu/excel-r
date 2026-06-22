"""Limit table data-row reads for extraction and formula compile."""
import os
from typing import Any, Dict, List, Optional, Set

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

TABLE_DATA_ROW_LIMIT = int(os.getenv("TABLE_DATA_ROW_LIMIT", "10"))


def limit_data_rows(data_rows: List[int], limit: Optional[int] = None) -> List[int]:
    """Return at most `limit` data rows; shorter tables are unchanged."""
    cap = TABLE_DATA_ROW_LIMIT if limit is None else limit
    if not data_rows or len(data_rows) <= cap:
        return list(data_rows)
    return data_rows[:cap]


def compile_range_for_table(table: Dict[str, Any], limit: Optional[int] = None) -> str:
    """
    Shrink a detected table range to header/title/total/check rows plus the
    first N data rows. Tables with <= N data rows keep their full range.
    """
    table_range = table.get("table_range", "")
    if not table_range:
        return ""

    cap = TABLE_DATA_ROW_LIMIT if limit is None else limit
    row_classification = table.get("row_classification", {})
    data_rows = row_classification.get("data_rows", [])
    if not data_rows or len(data_rows) <= cap:
        return table_range

    rows_needed = set()
    for key in ("title_rows", "header_rows", "total_rows", "check_rows"):
        rows_needed.update(row_classification.get(key, []))
    rows_needed.update(data_rows[:cap])

    col_start = table.get("col_start")
    col_end = table.get("col_end")
    if col_start is None or col_end is None:
        return table_range

    min_row = min(rows_needed)
    max_row = max(rows_needed)
    return f"{get_column_letter(col_start)}{min_row}:{get_column_letter(col_end)}{max_row}"


def formulas_range_ref(file_path: str, sheet_name: str, range_a1: str) -> str:
    """
    Build a formulas-library range reference.

    Format required by formulas.ExcelModel.from_ranges:
        '[path/to/workbook.xlsx]SheetName'!A6:I16
    """
    path = os.path.normpath(os.path.abspath(file_path))
    safe_sheet = sheet_name.replace("'", "''")
    return f"'[{path}]{safe_sheet}'!{range_a1}"


def collect_compile_range_refs(
    file_path: str,
    detected_tables: List[Dict[str, Any]],
    summary_sheets: Set[str],
) -> List[str]:
    """Collect limited table range refs for formulas-library compile."""
    refs: List[str] = []
    seen: Set[str] = set()
    for table in detected_tables:
        sheet = table.get("sheet_name", "")
        tr = compile_range_for_table(table)
        if not sheet or not tr or sheet not in summary_sheets:
            continue
        ref = formulas_range_ref(file_path, sheet, tr)
        if ref not in seen:
            seen.add(ref)
            refs.append(ref)
    return refs


def count_formulas_in_scoped_tables(
    file_path: str,
    detected_tables: List[Dict[str, Any]],
    summary_sheets: Set[str],
) -> int:
    """Count formula cells inside limited table compile ranges only."""
    if not detected_tables or not summary_sheets:
        return 0

    count = 0
    wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    try:
        sheet_map = {ws.title: ws for ws in wb.worksheets}
        for table in detected_tables:
            sheet = table.get("sheet_name", "")
            if sheet not in summary_sheets or sheet not in sheet_map:
                continue
            tr = compile_range_for_table(table)
            if not tr:
                continue
            try:
                min_col, min_row, max_col, max_row = range_boundaries(tr)
            except ValueError:
                continue
            ws = sheet_map[sheet]
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    if cell.value and str(cell.value).startswith("="):
                        count += 1
    finally:
        wb.close()
    return count
