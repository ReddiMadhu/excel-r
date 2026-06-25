"""
Formula Parser — Hybrid formula parsing using `formulas` library + custom fallback.

Primary: Uses the `formulas` library to compile the workbook into a dependency graph,
supporting all 500+ Excel functions automatically.

Fallback: If `formulas` fails on a specific cell, falls back to the existing custom
parsing logic for SUMIFS, COUNTIFS, SUM, and arithmetic formulas.
"""
import os
import re
import time
import tempfile
import traceback
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from typing import Any, Dict, Iterator, List, Optional, Set

from src.utils.table_row_limit import (
    TABLE_DATA_ROW_LIMIT,
    collect_compile_range_refs,
    compile_range_for_table,
    count_formulas_in_scoped_tables,
)
from src.utils.timing_log import log_step

# ─────────────────────────────────────────────────────────────────
# formulas library integration
# ─────────────────────────────────────────────────────────────────

def shrink_formula_ranges(formula_str: str, limit: Optional[int] = None) -> str:
    """
    Rewrite full column references (e.g. A:A) and large ranges to capped A1 ranges
    to prevent formulas library from expanding them to 1M+ rows.

    Uses TABLE_DATA_ROW_LIMIT (default 10) unless `limit` is provided.
    """
    if not formula_str or not str(formula_str).startswith('='):
        return formula_str

    cap = TABLE_DATA_ROW_LIMIT if limit is None else limit

    # 1. Replace full column references like Sheet!A:B or Sheet!$A:$B
    def repl_col_range(match):
        sheet = match.group(1) or ""
        col1 = match.group(2)
        col2 = match.group(3)
        sheet_prefix = f"'{sheet}'!" if sheet else ""
        return f"{sheet_prefix}{col1}1:{col2}{cap}"

    pattern_col = r"(?:'?([^'!]+)'?!)?\$?([A-Z]+):\$?([A-Z]+)"
    shrunk = re.sub(pattern_col, repl_col_range, formula_str)

    # 2. Replace large cell ranges like Sheet!A1:B5000
    def repl_cell_range(match):
        sheet = match.group(1) or ""
        col1 = match.group(2)
        row1 = match.group(3)
        col2 = match.group(4)
        row2 = match.group(5)
        sheet_prefix = f"'{sheet}'!" if sheet else ""
        try:
            start_row = int(row1)
            end_row = int(row2)
            if end_row - start_row + 1 > cap:
                return f"{sheet_prefix}{col1}{row1}:{col2}{start_row + cap - 1}"
        except ValueError:
            pass
        return match.group(0)

    pattern_range = r"(?:'?([^'!]+)'?!)?\$?([A-Z]+)\$?([0-9]+):\$?([A-Z]+)\$?([0-9]+)"
    shrunk = re.sub(pattern_range, repl_cell_range, shrunk)

    return shrunk


def _safe_remove_file(path: str) -> None:
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except Exception as ex:
            print(f"Warning: Could not remove temporary file {path}: {ex}")


def _iter_formula_cells_in_scoped_tables(
    wb: openpyxl.Workbook,
    tables: List[Dict[str, Any]],
    summary_sheets: Set[str],
) -> Iterator[Any]:
    """Yield formula cells inside limited table compile ranges on summary sheets."""
    for table in tables:
        sheet = table.get("sheet_name", "")
        if sheet not in summary_sheets or sheet not in wb.sheetnames:
            continue
        table_range = compile_range_for_table(table)
        if not table_range:
            continue
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table_range)
        except ValueError:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    yield cell


def _create_shrunk_workbook_copy(
    file_path: str,
    summary_sheets: Set[str],
    tables: List[Dict[str, Any]],
) -> str:
    """
    Return a temp workbook path with shrunk formula ranges, or the original path on failure.
    """
    temp_dir = os.path.dirname(file_path)
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=temp_dir)
    os.close(fd)
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        for cell in _iter_formula_cells_in_scoped_tables(wb, tables, summary_sheets):
            shrunk_val = shrink_formula_ranges(str(cell.value))
            if shrunk_val != cell.value:
                cell.value = shrunk_val
        wb.save(temp_path)
        return temp_path
    except Exception as e:
        print(f"Warning: Failed to create/shrink temporary workbook: {e}")
        _safe_remove_file(temp_path)
        return file_path
    finally:
        if wb is not None:
            wb.close()


_node_lookup = None


def _count_formulas_in_sheets(file_path: str, sheet_names: Set[str]) -> int:
    """Count formula cells limited to the given sheet names."""
    if not sheet_names:
        return 0
    count = 0
    wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    try:
        for ws in wb.worksheets:
            if ws.title not in sheet_names:
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).startswith("="):
                        count += 1
    finally:
        wb.close()
    return count


def _summary_sheet_names(sheet_types: Optional[Dict[str, str]]) -> Set[str]:
    if not sheet_types:
        return set()
    return {name for name, stype in sheet_types.items() if stype == "summary_report"}


def _compile_full_summary_sheets(
    file_path: str,
    summary_sheets: Set[str],
    scoped_count: int,
) -> Any:
    """Legacy path: push entire summary sheets (FORMULAS_LIB_USE_TABLE_RANGES=full)."""
    t0 = time.perf_counter()
    try:
        import formulas

        model = formulas.ExcelModel()
        book, context = model.add_book(file_path)
        for ws in book.worksheets:
            if ws.title in summary_sheets:
                model.push(ws, context)
        model.finish(complete=False, assemble=True)
        elapsed = time.perf_counter() - t0
        log_step(
            "formulas_compile",
            "scoped_summary_sheets",
            elapsed,
            sheets=",".join(sorted(summary_sheets)),
            scoped_formulas=scoped_count,
            table_ranges=False,
            data_row_limit=TABLE_DATA_ROW_LIMIT,
            compile_mode="full_summary_push",
        )
        print(
            f"  [formulas] Full summary push: {len(summary_sheets)} sheet(s), "
            f"{scoped_count} formula(s) in scope"
        )
        return model
    except Exception as e:
        print(f"Warning: full summary formulas compile failed: {e}")
        traceback.print_exc()
        return None


def compile_workbook_scoped(
    file_path: str,
    sheet_types: Optional[Dict[str, str]] = None,
    detected_tables: Optional[List[Dict[str, Any]]] = None,
    max_cells: int = 5000,
) -> Any:
    """
    Compile a reduced dependency graph — summary/report sheets only by default.

    Cross-sheet references (e.g. SUMIFS on SQL_data) are registered on summary
    formulas; raw data sheets are not fully compiled unless lazily expanded.

    Set FORMULAS_LIB_SCOPE=full to restore whole-workbook compile.
    Set FORMULAS_LIB_USE_TABLE_RANGES=full to compile whole summary sheets.
    Table compile ranges include at most TABLE_DATA_ROW_LIMIT (default 10) data rows.
    Set max_cells=0 to disable.
    """
    if max_cells <= 0:
        return None

    scope = os.getenv("FORMULAS_LIB_SCOPE", "scoped").strip().lower()
    if scope == "full":
        return compile_workbook_with_budget(file_path, max_cells=max_cells)

    summary_sheets = _summary_sheet_names(sheet_types)
    if not summary_sheets:
        print("Warning: no summary_report sheets — falling back to full formulas compile")
        return compile_workbook_with_budget(file_path, max_cells=max_cells)

    use_full_sheet = os.getenv("FORMULAS_LIB_USE_TABLE_RANGES", "").strip().lower() in (
        "0", "false", "no", "full",
    )
    if use_full_sheet:
        scoped_count = _count_formulas_in_sheets(file_path, summary_sheets)
        if scoped_count > max_cells:
            print(
                f"Warning: scoped formulas library skipped — "
                f"{scoped_count} summary formulas exceeds budget {max_cells}"
            )
            return None
        return _compile_full_summary_sheets(file_path, summary_sheets, scoped_count)

    tables = detected_tables or []
    scoped_count = count_formulas_in_scoped_tables(file_path, tables, summary_sheets)
    if scoped_count == 0:
        log_step(
            "formulas_compile",
            "scoped_summary_sheets",
            0.0,
            sheets=",".join(sorted(summary_sheets)),
            scoped_formulas=0,
            table_ranges=False,
            data_row_limit=TABLE_DATA_ROW_LIMIT,
            compile_skipped="no_formulas",
        )
        print("  [formulas] Skipped compile — no formulas in scoped table ranges")
        return None

    if scoped_count > max_cells:
        print(
            f"Warning: scoped formulas library skipped — "
            f"{scoped_count} scoped formulas exceeds budget {max_cells}"
        )
        return None

    compile_file_path = _create_shrunk_workbook_copy(file_path, summary_sheets, tables)
    using_temp = compile_file_path != file_path

    try:
        range_refs = collect_compile_range_refs(compile_file_path, tables, summary_sheets)
        if not range_refs:
            log_step(
                "formulas_compile",
                "scoped_summary_sheets",
                0.0,
                sheets=",".join(sorted(summary_sheets)),
                scoped_formulas=scoped_count,
                table_ranges=False,
                data_row_limit=TABLE_DATA_ROW_LIMIT,
                compile_skipped="no_table_ranges",
            )
            print("  [formulas] Skipped compile — no table ranges for summary sheets")
            return None

        t0 = time.perf_counter()
        try:
            import formulas

            model = formulas.ExcelModel()
            model.from_ranges(*range_refs)
            model.finish(complete=False, assemble=True)
            elapsed = time.perf_counter() - t0
            sample_ref = range_refs[0]
            if len(sample_ref) > 80:
                sample_ref = sample_ref[:77] + "..."
            log_step(
                "formulas_compile",
                "scoped_summary_sheets",
                elapsed,
                sheets=",".join(sorted(summary_sheets)),
                scoped_formulas=scoped_count,
                table_ranges=True,
                data_row_limit=TABLE_DATA_ROW_LIMIT,
                range_ref_count=len(range_refs),
                range_refs_sample=sample_ref,
            )
            print(
                f"  [formulas] Scoped compile: {len(summary_sheets)} summary sheet(s), "
                f"{scoped_count} formula(s), {len(range_refs)} range(s), "
                f"{TABLE_DATA_ROW_LIMIT} data rows max"
            )
            return model
        except Exception as e:
            elapsed = time.perf_counter() - t0
            print(f"Warning: table-range compile failed: {e}")
            traceback.print_exc()
            log_step(
                "formulas_compile",
                "scoped_summary_sheets",
                elapsed,
                sheets=",".join(sorted(summary_sheets)),
                scoped_formulas=scoped_count,
                table_ranges=False,
                data_row_limit=TABLE_DATA_ROW_LIMIT,
                compile_skipped="from_ranges_failed",
                range_ref_count=len(range_refs),
            )
            return None
    finally:
        if using_temp:
            _safe_remove_file(compile_file_path)


def compile_workbook(file_path):
    """
    Compile an entire workbook using the `formulas` library.
    Returns the ExcelModel or None if compilation fails.
    """
    return compile_workbook_with_budget(file_path)


def compile_workbook_with_budget(file_path, max_cells=5000):
    """
    Compile workbook with formulas library, respecting a cell/formula budget.

    Skips compilation if the workbook exceeds max_cells estimated formulas
    to avoid OOM. Set max_cells=0 to disable entirely.
    """
    if max_cells <= 0:
        return None
    try:
        import formulas
        import openpyxl

        # Quick estimate: count formula cells before full compile
        formula_count = 0
        wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).startswith("="):
                        formula_count += 1
                        if formula_count > max_cells:
                            wb.close()
                            print(
                                f"Warning: formulas library skipped — "
                                f"{formula_count} formulas exceeds budget {max_cells}"
                            )
                            return None
        wb.close()

        xl_model = formulas.ExcelModel().loads(file_path).finish()
        return xl_model
    except Exception as e:
        print(f"Warning: formulas library compilation failed: {e}")
        return None


def get_cell_dependencies(xl_model, sheet_name, cell_ref):
    """
    Get all cells that a given cell depends on (predecessors) using the
    formulas library dependency graph.
    
    Returns a list of cell reference strings like "'[wb]Sheet!A1'".
    """
    global _node_lookup
    if xl_model is None:
        return []
    
    try:
        dsp = xl_model.dsp
        
        # Build lookup if not cached or model changed
        if _node_lookup is None or _node_lookup.get("dsp_id") != id(dsp):
            lookup = {}
            for node in dsp.nodes:
                node_str = str(node)
                if '!' in node_str:
                    parts = node_str.split('!')
                    sheet_part = parts[0].strip("'\"[]")
                    cell_part = parts[1].strip("'\"$")
                    
                    # Extract sheet name (remove workbook prefix)
                    if ']' in sheet_part:
                        s_name = sheet_part.split(']')[-1]
                    else:
                        s_name = sheet_part
                    
                    key = (s_name.upper(), cell_part.upper())
                    lookup[key] = node
            _node_lookup = {
                "dsp_id": id(dsp),
                "lookup": lookup
            }
            
        cell_upper = cell_ref.upper().replace('$', '')
        target_key = _node_lookup["lookup"].get((sheet_name.upper(), cell_upper))
        
        if target_key is None:
            return []
        
        # Get all predecessors (cells this cell depends on)
        predecessors = []
        try:
            preds = list(dsp.predecessors(target_key))
            predecessors = [str(p) for p in preds]
        except Exception:
            pass
        
        return predecessors
    except Exception:
        return []


def resolve_dependency_to_column(dep_ref, raw_column_maps):
    """
    Translate a formulas-library dependency reference back to a human-readable
    column name using the raw column maps.
    
    dep_ref format: "'[workbook.xlsx]SheetName'!A1" or similar
    """
    try:
        # Extract sheet name and cell ref
        if '!' in dep_ref:
            parts = dep_ref.split('!')
            sheet_part = parts[0].strip("'\"[]")
            cell_part = parts[1].strip("'\"")
            
            # Extract just the sheet name (remove workbook name)
            if ']' in sheet_part:
                sheet_name = sheet_part.split(']')[-1]
            else:
                sheet_name = sheet_part
            
            # Get column letter from cell reference
            col_letter = ''.join(filter(str.isalpha, cell_part)).upper()
            
            # Look up in raw column maps
            if sheet_name in raw_column_maps:
                header = raw_column_maps[sheet_name].get(col_letter)
                if header:
                    return {
                        "sheet": sheet_name,
                        "column": col_letter,
                        "header": header,
                    }
            
            return {
                "sheet": sheet_name,
                "column": col_letter,
                "header": None,
            }
    except Exception:
        pass
    
    return None


def _parse_formula_with_library_impl(xl_model, sheet_name, cell_ref, formula_str,
                                     current_row, ws_val, raw_column_maps, 
                                     table_col_mapping, table_name, ws_form=None,
                                     detected_tables=None):
    """
    Parse a formula using the formulas library dependency graph.
    Falls back to custom parsing if the library can't resolve it.
    """
    deps = get_cell_dependencies(xl_model, sheet_name, cell_ref)
    
    if not deps:
        # Fallback to custom parsing
        return parse_formula(
            formula_str, current_row, ws_val, 
            raw_column_maps, table_col_mapping, table_name,
            ws_form,
            detected_tables=detected_tables
        )
    
    # Resolve dependencies to column names
    data_source_sheet = ""
    data_source_columns = set()
    formula_source_details = []
    
    for dep in deps:
        resolved = resolve_dependency_to_column(dep, raw_column_maps)
        if resolved and resolved["header"]:
            data_source_sheet = resolved["sheet"]
            data_source_columns.add(resolved["header"])
            formula_source_details.append({
                "column_name": resolved["header"],
                "role": "dependency",
            })
    
    if data_source_columns:
        # Successfully resolved via library
        # Generate a human-readable formula pattern
        formula_pattern = _generate_pattern_from_formula(
            formula_str, ws_val, current_row, raw_column_maps, table_col_mapping
        )
        
        return {
            "type": "formula_based",
            "formula_pattern": formula_pattern,
            "data_source_sheet": data_source_sheet,
            "data_source_columns": list(data_source_columns),
            "formula_source_details": formula_source_details,
            "formula_count": 1,
            "resolved_by": "formulas_library",
        }
    
    # Library found dependencies but couldn't map to raw columns
    # Fall back to custom parsing
    return parse_formula(
        formula_str, current_row, ws_val,
        raw_column_maps, table_col_mapping, table_name,
        ws_form,
        detected_tables=detected_tables
    )


def _clean_criteria_value(crit_val_repr):
    """
    Extract a clean, human-readable value from a criteria value representation.
    
    Examples:
      "Summary!K5 ('Flexible')" -> "'Flexible'"
      "current RBC C2 Product Category" -> "current row"
      "Col_A" -> "Col_A"
    """
    if "current " in crit_val_repr:
        return "current row"
    # Look for a value in parentheses e.g. ('Flexible') or ("Flexible")
    match = re.search(r"\(['\"](.+?)['\"]\)", crit_val_repr)
    if match:
        return f"'{match.group(1)}'"
    # Fall back to stripping quotes
    return crit_val_repr.strip("\"'")


def _build_sql_pattern(func_name, sum_repr, filters, criterias, is_negative=False):
    """
    Build a standardized SQL-like formula pattern string.
    
    Output format:
      [-1 * ] FUNC(source[column])
      [WHERE [col] = 'value' AND ...]
      [GROUP BY [col1], [col2]]
    """
    pattern = f"{func_name}({sum_repr})"
    if filters:
        where_parts = []
        for f in filters:
            # f is already in format "ColumnName = value"
            where_parts.append(f"[{f}]")
        pattern += " WHERE " + " AND ".join(where_parts)
    if criterias:
        group_parts = [f"[{c}]" for c in criterias]
        pattern += " GROUP BY " + ", ".join(group_parts)
    if is_negative:
        pattern = f"-1 * {pattern}"
    return pattern


def _generate_pattern_from_formula(formula_str, ws_val, current_row,
                                    raw_column_maps, table_col_mapping):
    """
    Generate a human-readable SQL-like formula pattern from a raw formula string.
    Replaces cell references with header names, respecting function structure.
    """
    if not formula_str:
        return ""

    pattern = formula_str.lstrip('=')

    # Replace sheet!column:column references (e.g., SQL_data!D:D)
    sheet_col_refs = re.findall(
        r"'?([A-Za-z_][A-Za-z0-9_ ]*)'?!\$?([A-Z]+):\$?([A-Z]+)", pattern
    )
    for sheet, col1, col2 in sheet_col_refs:
        if sheet in raw_column_maps:
            header = raw_column_maps[sheet].get(col1.upper(), f"Col_{col1}")
            pattern = pattern.replace(f"'{sheet}'!${col1}:${col2}", f"{sheet}[{header}]")
            pattern = pattern.replace(f"'{sheet}'!{col1}:{col2}", f"{sheet}[{header}]")
            pattern = pattern.replace(f"{sheet}!${col1}:${col2}", f"{sheet}[{header}]")
            pattern = pattern.replace(f"{sheet}!{col1}:{col2}", f"{sheet}[{header}]")

    # Replace cell references (with optional sheet names)
    cell_refs = re.findall(r"(?:(?:'[^']+'|[A-Za-z0-9_\-]+)!)?\$?[A-Z]+\$?[0-9]+", pattern)
    for ref in set(cell_refs):
        repr_val, header, ref_sheet = translate_reference(
            ref, current_row, ws_val, raw_column_maps, table_col_mapping, detected_tables=None
        )
        pattern = re.sub(
            rf'(?<![A-Za-z0-9_]){re.escape(ref)}(?![A-Za-z0-9_])',
            repr_val,
            pattern
        )

    return pattern


# ─────────────────────────────────────────────────────────────────
# Custom formula parsing (fallback)
# These are kept from the original code as a reliable fallback.
# ─────────────────────────────────────────────────────────────────

def is_value_cell(cell_val, cell_formula):
    """Determine if a cell represents a value or formula (as opposed to metadata/labels)."""
    if cell_val is None:
        return False
    if str(cell_formula).startswith('='):
        return True
    if isinstance(cell_val, (int, float)):
        return True
    return False


def extract_function_calls(formula_str, func_name):
    """
    Find all occurrences of a function call (like SUMIFS, COUNTIFS, SUM) 
    in the formula and extract their arguments, respecting nested parentheses.
    """
    calls = []
    idx = 0
    formula_upper = formula_str.upper()
    func_pattern = func_name.upper() + "("
    
    while True:
        pos = formula_upper.find(func_pattern, idx)
        if pos == -1:
            break
            
        start_arg = pos + len(func_pattern)
        paren_count = 1
        curr = start_arg
        
        while curr < len(formula_str) and paren_count > 0:
            char = formula_str[curr]
            if char == '(':
                paren_count += 1
            elif char == ')':
                paren_count -= 1
            curr += 1
            
        if paren_count == 0:
            call_str = formula_str[pos:curr]
            args_str = formula_str[start_arg:curr-1]
            
            # Split arguments by comma, respecting nested parens and quotes
            args = []
            current_arg = []
            inner_paren = 0
            in_quote = False
            
            for char in args_str:
                if char == '"':
                    in_quote = not in_quote
                    current_arg.append(char)
                elif char == ',' and not in_quote and inner_paren == 0:
                    args.append("".join(current_arg).strip())
                    current_arg = []
                else:
                    if char == '(':
                        inner_paren += 1
                    elif char == ')':
                        inner_paren -= 1
                    current_arg.append(char)
            if current_arg:
                args.append("".join(current_arg).strip())
                
            calls.append({
                "full_call": call_str,
                "args": args
            })
            idx = curr
        else:
            idx = pos + len(func_pattern)
            
    return calls


def parse_sheet_cell_ref(ref_str):
    """
    Parse a reference string like SQL_data!D:D or 'Warehouse Data'!$C:$C or Summary!$A$2 or B6.
    Returns (sheet, col_letter, row_num, is_range) or (None, None, None, False).
    """
    # Remove absolute signs
    clean_ref = ref_str.replace('$', '')
    
    sheet_name = None
    cell_part = clean_ref
    
    if '!' in clean_ref:
        parts = clean_ref.split('!')
        sheet_name = parts[0].strip("'")
        cell_part = parts[1]
        
    # Check if it is a column range e.g. D:D or C:C
    if ':' in cell_part:
        sub_parts = cell_part.split(':')
        if sub_parts[0].isalpha() and sub_parts[1].isalpha():
            return sheet_name, sub_parts[0].upper(), None, True
        else:
            # Row range or mixed range, let's extract column letter of first cell
            col_letter = "".join(filter(str.isalpha, sub_parts[0]))
            return sheet_name, col_letter.upper(), None, True
            
    # Single cell ref e.g. A2 or B6
    col_letter = "".join(filter(str.isalpha, cell_part))
    row_num_str = "".join(filter(str.isdigit, cell_part))
    row_num = int(row_num_str) if row_num_str else None
    
    return sheet_name, col_letter.upper(), row_num, False


def translate_reference(ref_str, current_row, summary_ws_val, raw_column_maps, table_col_mapping, detected_tables=None):
    """
    Translate a single cell/column reference in a formula to its header/value representation.
    """
    sheet_name, col_letter, row_num, is_range = parse_sheet_cell_ref(ref_str)
    
    current_sheet_title = summary_ws_val.title if summary_ws_val else "Summary"
    
    # If referencing a raw data sheet
    if sheet_name and sheet_name in raw_column_maps:
        col_map = raw_column_maps[sheet_name]
        header = col_map.get(col_letter, f"Col_{col_letter}")
        return f"{sheet_name}[{header}]", header, sheet_name
        
    # Resolve the target worksheet dynamically
    wb = summary_ws_val.parent if summary_ws_val else None
    ref_ws = summary_ws_val
    ref_sheet_title = current_sheet_title
    is_valid_ref = False
    
    if not sheet_name:
        is_valid_ref = True
    elif wb and sheet_name in wb.sheetnames:
        ref_ws = wb[sheet_name]
        ref_sheet_title = sheet_name
        is_valid_ref = True
        
    if is_valid_ref and col_letter:
        col_idx = column_index_from_string(col_letter)
        
        # Get header name
        header = f"Col_{col_letter}"
        if ref_sheet_title == current_sheet_title:
            header = table_col_mapping.get(col_idx, f"Col_{col_letter}")
        elif ref_ws:
            h_val = ref_ws.cell(row=2, column=col_idx).value
            if not h_val:
                h_val = ref_ws.cell(row=1, column=col_idx).value
            if h_val:
                header = str(h_val).strip()
        
        # If we have detected_tables and we are referencing the current sheet,
        # try to map the cell to a specific table's column.
        if ref_sheet_title == current_sheet_title and detected_tables and row_num:
            for tbl in detected_tables:
                if tbl.get("row_start") and tbl.get("row_end") and tbl.get("col_start") and tbl.get("col_end"):
                    if tbl["row_start"] <= row_num <= tbl["row_end"] and tbl["col_start"] <= col_idx <= tbl["col_end"]:
                        col_offset = col_idx - tbl["col_start"]
                        headers = tbl.get("headers", [])
                        if col_offset < len(headers):
                            header_name = headers[col_offset]
                            return f"{tbl['table_name']}[{header_name}]", header_name, current_sheet_title
        
        # Fallback to local table col mapping on current sheet
        # If it refers to the same row as the formula on the current sheet
        if ref_sheet_title == current_sheet_title and row_num == current_row:
            return f"current {header}", header, current_sheet_title
            
        # If it refers to a specific cell on the referenced sheet
        if row_num and ref_ws:
            val = ref_ws.cell(row=row_num, column=col_idx).value
            
            if isinstance(val, (int, float)):
                val_str = str(val)
            else:
                val_str = f"'{val}'" if val is not None else "None"
                
            return f"{ref_sheet_title}!{col_letter}{row_num} ({val_str})", header, ref_sheet_title
            
    return ref_str, None, None


def _parse_formula_impl(formula_str, current_row, summary_ws_val, raw_column_maps, table_col_mapping, table_name, summary_ws_form=None, _depth=0, detected_tables=None):
    """
    Parse an Excel formula and extract metadata, lineage and formula patterns.
    
    This is the original custom parser kept as a fallback for when the 
    formulas library can't resolve a specific cell.
    
    _depth: recursion depth for transitive resolution (max 5 to prevent infinite loops).
    """
    if not formula_str or not str(formula_str).startswith('='):
        return {
            "type": "raw",
            "formula_pattern": "",
            "data_source_sheet": "",
            "data_source_columns": [],
            "formula_source_details": [],
            "formula_count": 0,
            "resolved_by": "none",
        }
        
    formula_upper = formula_str.upper()
    
    current_sheet_title = summary_ws_val.title if summary_ws_val else "Summary"
    
    data_source_sheet = ""
    data_source_columns = set()
    formula_source_details = []
    formula_pattern = formula_str
    
    # 1. Parse SUMIFS
    sumifs_calls = extract_function_calls(formula_str, "SUMIFS")
    if sumifs_calls:
        patterns = []
        for call in sumifs_calls:
            args = call["args"]
            if len(args) < 3:
                continue
                
            sum_ref = args[0]
            sum_repr, sum_hdr, sum_sheet = translate_reference(sum_ref, current_row, summary_ws_val, raw_column_maps, table_col_mapping, detected_tables)
            if sum_hdr:
                data_source_columns.add(sum_hdr)
                if sum_sheet:
                    data_source_sheet = sum_sheet
                formula_source_details.append({
                    "column_name": sum_hdr,
                    "role": "sum_range"
                })
                
            criterias = []
            filters = []
            for i in range(1, len(args), 2):
                if i + 1 >= len(args):
                    break
                crit_range_ref = args[i]
                crit_val_ref = args[i+1]

                crit_range_repr, crit_range_hdr, crit_sheet = translate_reference(
                    crit_range_ref, current_row, summary_ws_val,
                    raw_column_maps, table_col_mapping, detected_tables
                )
                crit_val_repr, crit_val_hdr, _ = translate_reference(
                    crit_val_ref, current_row, summary_ws_val,
                    raw_column_maps, table_col_mapping, detected_tables
                )

                if crit_range_hdr:
                    data_source_columns.add(crit_range_hdr)
                    if crit_sheet:
                        data_source_sheet = crit_sheet

                col_label = crit_range_hdr if crit_range_hdr else crit_range_repr

                # Classify using $ pattern — the definitive way to detect GROUP BY vs WHERE
                from src.parsers.formula_lineage import classify_criteria_ref, extract_filter_value
                ref_type = classify_criteria_ref(crit_val_ref, current_row)

                if ref_type == 'group_by_key':
                    criterias.append(col_label)
                    formula_source_details.append({
                        "column_name": col_label,
                        "role": "group_by_key"
                    })
                else:
                    fv = extract_filter_value(crit_val_ref, crit_val_repr, summary_ws_val)
                    filters.append(f"{col_label} = {fv}")
                    formula_source_details.append({
                        "column_name": col_label,
                        "role": "static_filter",
                        "filter_value": fv
                    })

            is_negative = formula_str.strip().startswith('=-1*') or formula_str.strip().startswith('=-')
            pattern = _build_sql_pattern("SUM", sum_repr, filters, criterias)
            formula_pattern = formula_pattern.replace(call["full_call"], pattern)

        clean_pattern = formula_pattern.lstrip('=-1*').lstrip('=-').lstrip('=')

        return {
            "type": "formula_based",
            "formula_pattern": f"-1 * {clean_pattern}" if is_negative else clean_pattern,
            "data_source_sheet": data_source_sheet,
            "data_source_columns": list(data_source_columns),
            "formula_source_details": formula_source_details,
            "formula_count": len(sumifs_calls),
            "resolved_by": "custom_parser",
        }
        
    # 2. Parse COUNTIFS
    countifs_calls = extract_function_calls(formula_str, "COUNTIFS")
    if countifs_calls:
        for call in countifs_calls:
            args = call["args"]
            if len(args) < 2:
                continue

            criterias = []
            filters = []
            for i in range(0, len(args), 2):
                if i + 1 >= len(args):
                    break
                crit_range_ref = args[i]
                crit_val_ref = args[i+1]

                crit_range_repr, crit_range_hdr, crit_sheet = translate_reference(
                    crit_range_ref, current_row, summary_ws_val,
                    raw_column_maps, table_col_mapping, detected_tables
                )
                crit_val_repr, crit_val_hdr, _ = translate_reference(
                    crit_val_ref, current_row, summary_ws_val,
                    raw_column_maps, table_col_mapping, detected_tables
                )

                if crit_range_hdr:
                    data_source_columns.add(crit_range_hdr)
                    if crit_sheet:
                        data_source_sheet = crit_sheet

                col_label = crit_range_hdr if crit_range_hdr else crit_range_repr

                from src.parsers.formula_lineage import classify_criteria_ref, extract_filter_value
                ref_type = classify_criteria_ref(crit_val_ref, current_row)

                if ref_type == 'group_by_key':
                    criterias.append(col_label)
                    formula_source_details.append({
                        "column_name": col_label,
                        "role": "group_by_key"
                    })
                else:
                    fv = extract_filter_value(crit_val_ref, crit_val_repr, summary_ws_val)
                    filters.append(f"{col_label} = {fv}")
                    formula_source_details.append({
                        "column_name": col_label,
                        "role": "static_filter",
                        "filter_value": fv
                    })

            count_source = data_source_sheet if data_source_sheet else "source"
            pattern = _build_sql_pattern("COUNT", count_source, filters, criterias)
            formula_pattern = formula_pattern.replace(call["full_call"], pattern)

        clean_pattern = formula_pattern.lstrip('=')
        return {
            "type": "formula_based",
            "formula_pattern": clean_pattern,
            "data_source_sheet": data_source_sheet,
            "data_source_columns": list(data_source_columns),
            "formula_source_details": formula_source_details,
            "formula_count": len(countifs_calls),
            "resolved_by": "custom_parser",
        }
        
    # 3. Parse SUM totals
    sum_calls = extract_function_calls(formula_str, "SUM")
    if sum_calls:
        sum_data_source_sheet = ""
        sum_data_source_columns = set()
        sum_formula_source_details = []
        
        for call in sum_calls:
            args = call["args"]
            if not args:
                continue
            arg_reprs = []
            for arg in args:
                if ':' in arg:
                    sub_parts = arg.split(':')
                    sheet_name, start_col, start_row_num, _ = parse_sheet_cell_ref(sub_parts[0])
                    _, end_col, end_row_num, _ = parse_sheet_cell_ref(sub_parts[1])
                    if start_col and end_col:
                        start_idx = column_index_from_string(start_col)
                        end_idx = column_index_from_string(end_col)
                        if start_idx > end_idx:
                            start_idx, end_idx = end_idx, start_idx
                            
                        for c_idx in range(start_idx, end_idx + 1):
                            c_letter = get_column_letter(c_idx)
                            header = table_col_mapping.get(c_idx, f"Col_{c_letter}")
                            arg_reprs.append(f"SUM({header})")
                            
                            if not sheet_name or sheet_name == current_sheet_title:
                                sum_data_source_columns.add(header)
                                if not sum_data_source_sheet or sum_data_source_sheet == current_sheet_title:
                                    sum_data_source_sheet = current_sheet_title
                        
                        # Transitive resolution: resolve cells in the range up to a safe limit (50)
                        if summary_ws_form and _depth < 5 and start_row_num and end_row_num and (not sheet_name or sheet_name == current_sheet_title):
                            cells_processed = 0
                            limit = 50
                            break_outer = False
                            for r_idx in range(start_row_num, end_row_num + 1):
                                if break_outer:
                                    break
                                for c_idx in range(start_idx, end_idx + 1):
                                    cells_processed += 1
                                    if cells_processed > limit:
                                        break_outer = True
                                        break
                                    ref_formula = summary_ws_form.cell(row=r_idx, column=c_idx).value
                                    if ref_formula and str(ref_formula).startswith('='):
                                        parsed_ref = parse_formula(
                                            ref_formula, r_idx, summary_ws_val,
                                            raw_column_maps, table_col_mapping, table_name,
                                            summary_ws_form, _depth=_depth + 1,
                                            detected_tables=detected_tables
                                        )
                                        if parsed_ref["data_source_columns"]:
                                            if parsed_ref["data_source_sheet"] and parsed_ref["data_source_sheet"] != current_sheet_title:
                                                if sum_data_source_sheet == current_sheet_title:
                                                    sum_data_source_columns = set()
                                                sum_data_source_sheet = parsed_ref["data_source_sheet"]
                                                sum_data_source_columns.update(parsed_ref["data_source_columns"])
                                            else:
                                                if not sum_data_source_sheet or sum_data_source_sheet == current_sheet_title:
                                                    sum_data_source_sheet = current_sheet_title
                                                    sum_data_source_columns.update(parsed_ref["data_source_columns"])
                else:
                    repr_val, header, ref_sheet = translate_reference(arg, current_row, summary_ws_val, raw_column_maps, table_col_mapping, detected_tables)
                    arg_reprs.append(repr_val)
                    if header:
                        sum_formula_source_details.append({
                            "column_name": header,
                            "role": "sum_input"
                        })
                        if ref_sheet == current_sheet_title:
                            sum_data_source_columns.add(header)
                            if not sum_data_source_sheet or sum_data_source_sheet == current_sheet_title:
                                sum_data_source_sheet = current_sheet_title
                        elif ref_sheet:
                            if sum_data_source_sheet == current_sheet_title:
                                sum_data_source_columns = set()
                            sum_data_source_sheet = ref_sheet
                            sum_data_source_columns.add(header)
                            
            formula_pattern = formula_pattern.replace(call["full_call"], " + ".join(arg_reprs))
            
        clean_pattern = formula_pattern.lstrip('=')
        
        # Translate any remaining cell references (like F7) in the clean pattern
        cell_refs = re.findall(r"(?:(?:'[^']+'|[A-Za-z0-9_\-]+)!)?\$?[A-Z]+\$?[0-9]+", clean_pattern)
        for ref in set(cell_refs):
            repr_val, header, ref_sheet = translate_reference(ref, current_row, summary_ws_val, raw_column_maps, table_col_mapping, detected_tables)
            clean_pattern = re.sub(rf'(?<![A-Za-z0-9_]){re.escape(ref)}(?![A-Za-z0-9_])', repr_val, clean_pattern)
            if header:
                if ref_sheet == current_sheet_title:
                    sum_data_source_columns.add(header)
                    if not sum_data_source_sheet or sum_data_source_sheet == current_sheet_title:
                        sum_data_source_sheet = current_sheet_title
                elif ref_sheet:
                    if sum_data_source_sheet == current_sheet_title:
                        sum_data_source_columns = set()
                    sum_data_source_sheet = ref_sheet
                    sum_data_source_columns.add(header)
            
        # If we found transitive/calculated data sources, return as formula_based
        if sum_data_source_columns:
            return {
                "type": "formula_based",
                "formula_pattern": clean_pattern,
                "data_source_sheet": sum_data_source_sheet,
                "data_source_columns": list(sum_data_source_columns),
                "formula_source_details": sum_formula_source_details,
                "formula_count": len(sum_calls),
                "resolved_by": "transitive_custom_parser",
            }
        
        return {
            "type": "total",
            "formula_pattern": clean_pattern,
            "data_source_sheet": "",
            "data_source_columns": [],
            "formula_source_details": [],
            "formula_count": len(sum_calls),
            "resolved_by": "custom_parser",
        }
        
    # 4. Parse arithmetic formulas
    arith_pattern = formula_str.lstrip('=')
    
    cell_refs = re.findall(r"(?:(?:'[^']+'|[A-Za-z0-9_\-]+)!)?\$?[A-Z]+\$?[0-9]+", arith_pattern)
    for ref in set(cell_refs):
        repr_val, header, ref_sheet = translate_reference(ref, current_row, summary_ws_val, raw_column_maps, table_col_mapping, detected_tables)
        arith_pattern = re.sub(rf'(?<![A-Za-z0-9_]){re.escape(ref)}(?![A-Za-z0-9_])', repr_val, arith_pattern)
        if header:
            formula_source_details.append({
                "column_name": header,
                "role": "arithmetic_input"
            })
            if ref_sheet == current_sheet_title:
                data_source_columns.add(header)
                if not data_source_sheet or data_source_sheet == current_sheet_title:
                    data_source_sheet = current_sheet_title
            elif ref_sheet:
                if data_source_sheet == current_sheet_title:
                    data_source_columns = set()
                data_source_sheet = ref_sheet
                data_source_columns.add(header)
            
    # Transitive data sources resolution
    if summary_ws_form and _depth < 5:
        for ref in set(cell_refs):
            sheet_name, col_letter, row_num, is_range = parse_sheet_cell_ref(ref)
            if not sheet_name or sheet_name == current_sheet_title:
                if col_letter and row_num:
                    col_idx = column_index_from_string(col_letter)
                    ref_formula = summary_ws_form.cell(row=row_num, column=col_idx).value
                    if ref_formula and str(ref_formula).startswith('='):
                        # Parse recursively with incremented depth
                        parsed_ref = parse_formula(
                            ref_formula, row_num, summary_ws_val,
                            raw_column_maps, table_col_mapping, table_name,
                            summary_ws_form, _depth=_depth + 1,
                            detected_tables=detected_tables
                        )
                        if parsed_ref["data_source_columns"]:
                            if parsed_ref["data_source_sheet"] and parsed_ref["data_source_sheet"] != current_sheet_title:
                                if data_source_sheet == current_sheet_title:
                                    data_source_columns = set()
                                data_source_sheet = parsed_ref["data_source_sheet"]
                                data_source_columns.update(parsed_ref["data_source_columns"])
                            else:
                                if not data_source_sheet or data_source_sheet == current_sheet_title:
                                    data_source_sheet = current_sheet_title
                                    data_source_columns.update(parsed_ref["data_source_columns"])
            elif sheet_name in raw_column_maps:
                col_map = raw_column_maps[sheet_name]
                header = col_map.get(col_letter)
                if header:
                    if data_source_sheet == current_sheet_title:
                        data_source_columns = set()
                    data_source_columns.add(header)
                    data_source_sheet = sheet_name

    first_cell_val = str(summary_ws_val.cell(row=current_row, column=1).value).strip().lower()
    if "check" in first_cell_val:
        f_type = "check"
    elif "total" in first_cell_val or "grand total" in first_cell_val:
        f_type = "total"
    else:
        f_type = "formula_based"
        
    return {
        "type": f_type,
        "formula_pattern": arith_pattern,
        "data_source_sheet": data_source_sheet,
        "data_source_columns": list(data_source_columns),
        "formula_source_details": formula_source_details,
        "formula_count": 1,
        "resolved_by": "transitive_custom_parser" if data_source_columns else "custom_parser",
    }


def infer_formula_pattern_for_column(formulas_in_column):
    """
    Given a list of formulas in a summary column, return a single generalized formula pattern.
    """
    for f in formulas_in_column:
        if f and str(f).startswith('='):
            return f
    return ""


def calculate_formula_nesting_depth(formula_str):
    """
    Calculate the deepest nesting level of functions inside an Excel formula.
    Constants or simple references return 0.
    """
    if not formula_str or not str(formula_str).startswith('='):
        return 0
        
    formula = str(formula_str).lstrip('=').strip().upper()
    
    # Simple regex to find function calls like NAME(
    func_pattern = re.compile(r'\b([A-Z][A-Z0-9_\.]+)\s*\(')
    
    def get_inner_depth(expr):
        matches = list(func_pattern.finditer(expr))
        if not matches:
            return 0
            
        max_sub_depth = 0
        for match in matches:
            start_pos = match.end()
            # Extract parenthesized block by matching parentheses
            paren_count = 1
            end_pos = start_pos
            while paren_count > 0 and end_pos < len(expr):
                if expr[end_pos] == '(':
                    paren_count += 1
                elif expr[end_pos] == ')':
                    paren_count -= 1
                end_pos += 1
                
            arg_expr = expr[start_pos:end_pos-1]
            max_sub_depth = max(max_sub_depth, get_inner_depth(arg_expr))
            
        return 1 + max_sub_depth
        
    return get_inner_depth(formula)


def parse_formula(formula_str, *args, **kwargs):
    """
    Wrapper for _parse_formula_impl to inject nesting_depth and function_chain.
    """
    res = _parse_formula_impl(formula_str, *args, **kwargs)
    if isinstance(res, dict):
        res["nesting_depth"] = calculate_formula_nesting_depth(formula_str)
        formula_upper = str(formula_str).upper() if formula_str else ""
        res["function_chain"] = list(dict.fromkeys(re.findall(r'\b([A-Z][A-Z0-9_\.]+)\s*\(', formula_upper)))
    return res


def parse_formula_with_library(xl_model, sheet_name, cell_ref, formula_str, *args, **kwargs):
    """
    Wrapper for _parse_formula_with_library_impl to inject nesting_depth and function_chain.
    """
    res = _parse_formula_with_library_impl(xl_model, sheet_name, cell_ref, formula_str, *args, **kwargs)
    if isinstance(res, dict):
        res["nesting_depth"] = calculate_formula_nesting_depth(formula_str)
        formula_upper = str(formula_str).upper() if formula_str else ""
        res["function_chain"] = list(dict.fromkeys(re.findall(r'\b([A-Z][A-Z0-9_\.]+)\s*\(', formula_upper)))
    return res


def classify_summary_formula(formula_str, row_num, row_type, data_rows,
                              ws_val, ws_form, raw_column_maps,
                              table_col_mapping, table_name, col_idx,
                              detected_tables=None):
    """
    Classify a formula found in a total or check row and build lightweight lineage.

    Returns a dict with:
      - computation_type: SUM | SUMIFS | ARITHMETIC | VALIDATION | UNKNOWN
      - scope: column_total | filtered_total | cross_table_check | intra_table_check | arithmetic
      - references_data_rows: list of row numbers referenced (if applicable)
      - validates_against: source sheet/table name (for check rows)
      - validates_total_row: row number of the total row being checked (for check rows)
    """
    if not formula_str or not str(formula_str).startswith('='):
        return None

    formula_upper = formula_str.upper()
    current_sheet = ws_val.title if ws_val else "Summary"
    lineage = {}

    # ── Detect SUM of data rows (column_total) ──────────────────────
    sum_calls = extract_function_calls(formula_str, "SUM")
    sumifs_calls = extract_function_calls(formula_str, "SUMIFS")

    if sumifs_calls:
        # SUMIFS in a total/check row → filtered total or cross-table check
        args = sumifs_calls[0]["args"]
        sum_ref = args[0] if args else ""
        sheet_name, _, _, _ = parse_sheet_cell_ref(sum_ref)

        if row_type == "check":
            lineage["computation_type"] = "SUMIFS"
            lineage["scope"] = "cross_table_check"
            if sheet_name and sheet_name != current_sheet:
                lineage["validates_against"] = sheet_name
        else:
            lineage["computation_type"] = "SUMIFS"
            lineage["scope"] = "filtered_total"
            if sheet_name:
                lineage["source_sheet"] = sheet_name

    elif sum_calls:
        # Plain SUM — check if it references data rows in the same column
        args = sum_calls[0]["args"]
        referenced_rows = []
        references_external = False
        external_sheet = None

        for arg in args:
            if ':' in arg:
                parts = arg.split(':')
                sheet1, col1, row1, _ = parse_sheet_cell_ref(parts[0])
                _, col2, row2, _ = parse_sheet_cell_ref(parts[1])
                if sheet1 and sheet1 != current_sheet:
                    references_external = True
                    external_sheet = sheet1
                elif row1 and row2:
                    referenced_rows.extend(range(row1, row2 + 1))
            else:
                sheet_ref, _, row_ref, _ = parse_sheet_cell_ref(arg)
                if sheet_ref and sheet_ref != current_sheet:
                    references_external = True
                    external_sheet = sheet_ref
                elif row_ref:
                    referenced_rows.append(row_ref)

        # Check if the formula also contains arithmetic with external refs
        # e.g. =SUM(SQL_data!D:D)-B8
        remaining = formula_str
        for call in sum_calls:
            remaining = remaining.replace(call["full_call"], "")
        has_external_in_arithmetic = False
        ext_refs = re.findall(r"(?:'[^']+'|[A-Za-z0-9_\-]+)!\$?[A-Z]+", remaining)
        for ext_ref in ext_refs:
            ref_sheet = ext_ref.split('!')[0].strip("'")
            if ref_sheet != current_sheet:
                has_external_in_arithmetic = True
                external_sheet = ref_sheet

        if row_type == "check" and (references_external or has_external_in_arithmetic):
            lineage["computation_type"] = "VALIDATION"
            lineage["scope"] = "cross_table_check"
            if external_sheet:
                lineage["validates_against"] = external_sheet
            # Try to find total row reference in the formula
            cell_refs_in_formula = re.findall(r'\$?([A-Z]+)\$?(\d+)', formula_str)
            for _, rn in cell_refs_in_formula:
                rn_int = int(rn)
                if rn_int in (data_rows[-1] + 1 if data_rows else 0,) or rn_int == row_num - 1:
                    lineage["validates_total_row"] = rn_int
                    break
        elif row_type == "check" and not references_external:
            lineage["computation_type"] = "VALIDATION"
            lineage["scope"] = "intra_table_check"
            if referenced_rows:
                lineage["references_data_rows"] = [r for r in referenced_rows if r in data_rows]
        else:
            # Total row with SUM
            lineage["computation_type"] = "SUM"
            lineage["scope"] = "column_total"
            data_row_refs = [r for r in referenced_rows if r in data_rows]
            if data_row_refs:
                lineage["references_data_rows"] = data_row_refs

    else:
        # No SUM/SUMIFS — arithmetic formula
        # Check for external sheet references
        ext_refs = re.findall(r"(?:'[^']+'|[A-Za-z0-9_\-]+)!\$?[A-Z]+", formula_str)
        has_external = False
        external_sheet = None
        for ext_ref in ext_refs:
            ref_sheet = ext_ref.split('!')[0].strip("'")
            if ref_sheet != current_sheet:
                has_external = True
                external_sheet = ref_sheet

        # Extract all cell references to find data row refs
        cell_refs_in_formula = re.findall(r'\$?([A-Z]+)\$?(\d+)', formula_str)
        referenced_rows = []
        for _, rn in cell_refs_in_formula:
            referenced_rows.append(int(rn))

        if row_type == "check":
            if has_external:
                lineage["computation_type"] = "VALIDATION"
                lineage["scope"] = "cross_table_check"
                lineage["validates_against"] = external_sheet
            else:
                lineage["computation_type"] = "VALIDATION"
                lineage["scope"] = "intra_table_check"
        else:
            lineage["computation_type"] = "ARITHMETIC"
            lineage["scope"] = "column_total"
            data_row_refs = [r for r in referenced_rows if r in data_rows]
            if data_row_refs:
                lineage["references_data_rows"] = data_row_refs

    return lineage if lineage else None

