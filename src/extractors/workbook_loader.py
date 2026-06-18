"""
Workbook Loader — Load Excel workbooks and extract metadata.

Enhanced to also extract:
- Named ranges
- External links
- Print areas
- Hidden rows/columns
"""
import openpyxl
import hashlib
import os


def compute_md5(path):
    """Calculate MD5 hash of a file."""
    hash_md5 = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


class DummyCell:
    __slots__ = ['value', 'number_format', 'parent', 'font', 'border', 'fill', 'alignment', 'comment', 'hyperlink']
    def __init__(self, value=None, number_format="General", parent=None, font=None, border=None, fill=None, alignment=None, comment=None, hyperlink=None):
        self.value = value
        self.number_format = number_format
        self.parent = parent
        self.font = font
        self.border = border
        self.fill = fill
        self.alignment = alignment
        self.comment = comment
        self.hyperlink = hyperlink

class SheetCacheWrapper:
    def __init__(self, ws, cache_formatting=True):
        self._ws = ws
        self.title = ws.title
        self.min_row = ws.min_row or 1
        self.min_column = ws.min_column or 1
        self.max_row = ws.max_row or 1
        self.max_column = ws.max_column or 1
        
        self._data = []
        self._number_formats = [] if cache_formatting else None
        self._fonts = [] if cache_formatting else None
        self._borders = [] if cache_formatting else None
        self._fills = [] if cache_formatting else None
        self._alignments = [] if cache_formatting else None
        self._comments = [] if cache_formatting else None
        self._hyperlinks = [] if cache_formatting else None
        
        class DummyMergedCells:
            ranges = []
        self.merged_cells = getattr(ws, "merged_cells", DummyMergedCells())
        self.print_area = getattr(ws, "print_area", None)
        self.row_dimensions = getattr(ws, "row_dimensions", {})
        self.column_dimensions = getattr(ws, "column_dimensions", {})
        self.data_validations = getattr(ws, "data_validations", None)
        
        limit_rows = None
        if not cache_formatting:
            if ws.max_row is None or self.max_row > 200:
                limit_rows = 100
                
        print(f"  [Cache] Loading sheet '{ws.title}' into memory array...")
        row_count = 0
        for row in ws.iter_rows():
            self._data.append([c.value for c in row])
            if cache_formatting:
                self._number_formats.append([getattr(c, "number_format", "General") for c in row])
                self._fonts.append([getattr(c, "font", None) for c in row])
                self._borders.append([getattr(c, "border", None) for c in row])
                self._fills.append([getattr(c, "fill", None) for c in row])
                self._alignments.append([getattr(c, "alignment", None) for c in row])
                self._comments.append([getattr(c, "comment", None) for c in row])
                self._hyperlinks.append([getattr(c, "hyperlink", None) for c in row])
            row_count += 1
            if limit_rows and row_count >= limit_rows:
                break
            
        self.max_row = ws.max_row or row_count
        if self._data:
            self.max_column = max(self.max_column, max(len(r) for r in self._data))
            
    def cell(self, row, column):
        r_idx = row - 1
        c_idx = column - 1
        if 0 <= r_idx < len(self._data) and 0 <= c_idx < len(self._data[r_idx]):
            val = self._data[r_idx][c_idx]
            nf = self._number_formats[r_idx][c_idx] if self._number_formats else "General"
            font = self._fonts[r_idx][c_idx] if self._fonts else None
            border = self._borders[r_idx][c_idx] if self._borders else None
            fill = self._fills[r_idx][c_idx] if self._fills else None
            align = self._alignments[r_idx][c_idx] if self._alignments else None
            comment = self._comments[r_idx][c_idx] if self._comments else None
            hyperlink = self._hyperlinks[r_idx][c_idx] if self._hyperlinks else None
        else:
            val = None
            nf = "General"
            font = border = fill = align = comment = hyperlink = None
            
        return DummyCell(value=val, number_format=nf, parent=self, font=font, border=border, fill=fill, alignment=align, comment=comment, hyperlink=hyperlink)
        
    def iter_rows(self, min_row=1, max_row=None, min_col=1, max_col=None, values_only=False):
        if max_row is None:
            max_row = self.max_row
        if max_col is None:
            max_col = self.max_column
            
        for r_idx in range(min_row - 1, max_row):
            if r_idx >= len(self._data):
                break
            row_cells = []
            for c_idx in range(min_col - 1, max_col):
                if c_idx >= len(self._data[r_idx]):
                    if values_only:
                        row_cells.append(None)
                    else:
                        row_cells.append(self.cell(r_idx + 1, c_idx + 1))
                    continue
                if values_only:
                    row_cells.append(self._data[r_idx][c_idx])
                else:
                    row_cells.append(self.cell(r_idx + 1, c_idx + 1))
            yield tuple(row_cells)

    def __getattr__(self, item):
        return getattr(self._ws, item)

class WorkbookCacheWrapper:
    def __init__(self, wb, cache_formatting=True):
        self._wb = wb
        self._sheets = []
        self._sheet_dict = {}
        for ws in wb.worksheets:
            sheet_cache_fmt = cache_formatting
            if cache_formatting:
                name_lower = ws.title.lower()
                raw_keywords = ["data", "raw", "extract", "dump", "source", "query", "sql"]
                if any(kw in name_lower for kw in raw_keywords):
                    sheet_cache_fmt = False
                elif ws.max_row and ws.max_row > 500:
                    sheet_cache_fmt = False
                    
            cached_ws = SheetCacheWrapper(ws, cache_formatting=sheet_cache_fmt)
            cached_ws.parent = self
            self._sheets.append(cached_ws)
            self._sheet_dict[ws.title] = cached_ws
            
    @property
    def worksheets(self):
        return self._sheets
        
    @property
    def sheetnames(self):
        return [ws.title for ws in self._sheets]
        
    def __getitem__(self, key):
        return self._sheet_dict[key]

    def __getattr__(self, item):
        return getattr(self._wb, item)


def load_workbook_values(path):
    """Load workbook with data_only=True and read_only=True, then cache it."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return WorkbookCacheWrapper(wb, cache_formatting=True)


def load_workbook_formulas(path):
    """Load workbook with data_only=False and read_only=True, then cache it."""
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    return WorkbookCacheWrapper(wb, cache_formatting=False)


def get_sheet_used_range(ws):
    """Return the cell range of the sheet, e.g., 'A1:H50'."""
    min_col = ws.min_column
    min_row = ws.min_row
    max_col = ws.max_column
    max_row = ws.max_row
    if max_col is None or max_row is None or max_col < min_col or max_row < min_row:
        return "A1:A1"
    start_letter = openpyxl.utils.get_column_letter(min_col)
    end_letter = openpyxl.utils.get_column_letter(max_col)
    return f"{start_letter}{min_row}:{end_letter}{max_row}"


def get_non_empty_cells(ws):
    """Count non-empty cells in the sheet."""
    count = 0
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                count += 1
    return count


def get_formula_count(ws_formula):
    """Count number of cells containing formulas in the sheet."""
    count = 0
    for row in ws_formula.iter_rows(values_only=True):
        for val in row:
            if val is not None and str(val).startswith('='):
                count += 1
    return count


def extract_workbook_metadata(wb, file_path=None):
    """
    Extract workbook-level metadata that aids in understanding context.
    """
    meta = {
        "named_ranges": [],
        "external_links": [],
        "has_vba_macros": False,
        "vba_macro_streams": [],
    }
    
    # Extract VBA macros
    if file_path and os.path.exists(file_path):
        try:
            from oletools.olevba import VBA_Parser
            parser = VBA_Parser(file_path)
            if parser.detect_vba_macros():
                meta["has_vba_macros"] = True
                for sub_f, stream_path, vba_filename, vba_code in parser.extract_macros():
                    meta["vba_macro_streams"].append({
                        "stream_path": str(stream_path),
                        "vba_filename": str(vba_filename),
                        "code_length": len(vba_code),
                    })
        except Exception:
            pass

    # Named ranges
    try:
        if wb.defined_names:
            for name, dn in wb.defined_names.items():
                try:
                    meta["named_ranges"].append({
                        "name": name,
                        "value": dn.value,
                    })
                except Exception:
                    pass
    except Exception:
        pass
    
    # External links (references to other workbooks)
    try:
        if hasattr(wb, '_external_links') and wb._external_links:
            for link in wb._external_links:
                try:
                    if hasattr(link, 'file_link') and link.file_link:
                        meta["external_links"].append(str(link.file_link))
                    elif hasattr(link, 'Target'):
                        meta["external_links"].append(str(link.Target))
                except Exception:
                    pass
    except Exception:
        pass
    
    return meta



def extract_sheet_metadata(ws):
    """
    Extract sheet-level metadata including print area and hidden rows/columns.
    """
    meta = {
        "print_area": None,
        "hidden_rows": [],
        "hidden_columns": [],
    }
    
    # Print area — indicates what the user considers the "final output"
    try:
        if ws.print_area:
            meta["print_area"] = ws.print_area
    except Exception:
        pass
    
    # Hidden rows
    try:
        for row_idx, rd in ws.row_dimensions.items():
            if rd.hidden:
                meta["hidden_rows"].append(row_idx)
    except Exception:
        pass
    
    # Hidden columns
    try:
        for col_letter, cd in ws.column_dimensions.items():
            if cd.hidden:
                meta["hidden_columns"].append(col_letter)
    except Exception:
        pass
    
    return meta
