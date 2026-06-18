"""
Formatting Extractor — Extract visual/formatting metadata from Excel cells.

Humans understand Excel sheets by looking at formatting: bold = header, 
borders = section boundaries, indentation = hierarchy. This module makes 
that formatting data available to the parser.
"""
import openpyxl
from openpyxl.utils import get_column_letter


def extract_cell_formatting(cell):
    """
    Extract all formatting metadata from a single cell.
    Returns a dict of formatting attributes.
    """
    fmt = {
        "bold": False,
        "italic": False,
        "font_size": None,
        "font_color": None,
        "fill_color": None,
        "indent": 0,
        "border_bottom": None,
        "border_top": None,
        "border_right": None,
        "border_left": None,
        "number_format": "General",
        "has_comment": False,
        "comment_text": None,
        "hyperlink": None,
    }
    
    # Font
    if cell.font:
        fmt["bold"] = bool(cell.font.bold)
        fmt["italic"] = bool(cell.font.italic)
        fmt["font_size"] = cell.font.size
        if cell.font.color and cell.font.color.rgb and cell.font.color.rgb != "00000000":
            fmt["font_color"] = str(cell.font.color.rgb)
    
    # Fill / background color
    if cell.fill and cell.fill.fgColor:
        try:
            rgb = str(cell.fill.fgColor.rgb)
            if rgb and rgb != "00000000":
                fmt["fill_color"] = rgb
        except (AttributeError, TypeError):
            pass
    
    # Alignment / indentation
    if cell.alignment:
        fmt["indent"] = cell.alignment.indent or 0
    
    # Borders
    if cell.border:
        if cell.border.bottom and cell.border.bottom.style:
            fmt["border_bottom"] = cell.border.bottom.style
        if cell.border.top and cell.border.top.style:
            fmt["border_top"] = cell.border.top.style
        if cell.border.right and cell.border.right.style:
            fmt["border_right"] = cell.border.right.style
        if cell.border.left and cell.border.left.style:
            fmt["border_left"] = cell.border.left.style
    
    # Number format
    fmt["number_format"] = cell.number_format or "General"
    
    # Comments
    if cell.comment:
        fmt["has_comment"] = True
        fmt["comment_text"] = cell.comment.text
    
    # Hyperlinks
    if cell.hyperlink:
        fmt["hyperlink"] = cell.hyperlink.target
    
    return fmt


def analyze_row_formatting(ws, row_idx, col_start=1, col_end=None):
    """
    Analyze formatting signals for an entire row.
    Returns a summary of what the row 'looks like' to a human.
    """
    if col_end is None:
        col_end = ws.max_column or 1
    
    cells = []
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row_idx, column=c)
        cells.append(cell)
    
    non_empty = [c for c in cells if c.value is not None]
    if not non_empty:
        return {
            "is_blank": True,
            "is_header": False,
            "is_total": False,
            "is_section_title": False,
            "has_data": False,
            "bold_ratio": 0.0,
            "has_bottom_border": False,
            "max_indent": 0,
        }
    
    # Bold analysis
    bold_count = sum(1 for c in non_empty if c.font and c.font.bold)
    bold_ratio = bold_count / len(non_empty) if non_empty else 0
    
    # Border analysis
    has_bottom_border = any(
        c.border and c.border.bottom and c.border.bottom.style
        for c in non_empty
    )
    has_top_border = any(
        c.border and c.border.top and c.border.top.style
        for c in non_empty
    )
    
    # Indentation
    max_indent = max(
        (c.alignment.indent if c.alignment and c.alignment.indent else 0) 
        for c in non_empty
    )
    
    # Content analysis
    first_val = str(non_empty[0].value).strip().lower() if non_empty else ""
    has_numeric = any(isinstance(c.value, (int, float)) for c in non_empty)
    has_formula = any(
        c.value is not None and str(c.value).startswith('=') 
        for c in cells
    )
    
    # Classification
    is_header = (bold_ratio > 0.5 and has_bottom_border and not has_numeric)
    is_total = (
        ("total" in first_val or "grand total" in first_val)
        and (bold_ratio > 0.3 or has_bottom_border or has_top_border)
    )
    is_check = "check" in first_val
    is_section_title = (
        len(non_empty) == 1 
        and bold_ratio >= 1.0 
        and isinstance(non_empty[0].value, str)
    )
    has_data = has_numeric or has_formula
    
    return {
        "is_blank": False,
        "is_header": is_header,
        "is_total": is_total,
        "is_check": is_check,
        "is_section_title": is_section_title,
        "has_data": has_data,
        "bold_ratio": round(bold_ratio, 2),
        "has_bottom_border": has_bottom_border,
        "has_top_border": has_top_border,
        "max_indent": max_indent,
        "non_empty_count": len(non_empty),
        "first_value": first_val,
    }


def detect_row_hierarchy(ws, data_rows, label_col):
    """
    Build a parent-child hierarchy from cell indentation levels.
    
    Indent 0 = top level, indent 1 = child, indent 2 = grandchild, etc.
    Returns a list of dicts with row hierarchy information.
    """
    hierarchy = []
    parent_stack = []  # Stack of (indent_level, label)
    
    for row_idx in data_rows:
        cell = ws.cell(row=row_idx, column=label_col)
        indent = cell.alignment.indent if cell.alignment and cell.alignment.indent else 0
        label = str(cell.value).strip() if cell.value else ""
        
        if not label:
            # Keep last parent context for blank label cells
            parent = parent_stack[-1][1] if parent_stack else None
            hierarchy.append({
                "row": row_idx,
                "label": "",
                "indent_level": indent,
                "parent_label": parent,
                "full_path": " > ".join(p[1] for p in parent_stack) if parent_stack else "",
            })
            continue
        
        # Pop parents with same or deeper indent
        while parent_stack and parent_stack[-1][0] >= indent:
            parent_stack.pop()
        
        parent = parent_stack[-1][1] if parent_stack else None
        parent_stack.append((indent, label))
        
        hierarchy.append({
            "row": row_idx,
            "label": label,
            "indent_level": indent,
            "parent_label": parent,
            "full_path": " > ".join(p[1] for p in parent_stack),
        })
    
    return hierarchy


def extract_sheet_metadata(ws, wb=None):
    """
    Extract sheet-level metadata that aids in understanding context.
    """
    meta = {
        "merged_cells": [],
        "print_area": None,
        "hidden_rows": [],
        "hidden_columns": [],
        "named_ranges": [],
        "data_validations": [],
    }
    
    # Merged cells
    for rng in ws.merged_cells.ranges:
        meta["merged_cells"].append(str(rng))
    
    # Print area
    if ws.print_area:
        meta["print_area"] = ws.print_area
    
    # Hidden rows
    for row_idx, rd in ws.row_dimensions.items():
        if rd.hidden:
            meta["hidden_rows"].append(row_idx)
    
    # Hidden columns
    for col_letter, cd in ws.column_dimensions.items():
        if cd.hidden:
            meta["hidden_columns"].append(col_letter)
    
    # Data validations (dropdowns, etc.)
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            meta["data_validations"].append({
                "type": dv.type,
                "formula1": str(dv.formula1) if dv.formula1 else None,
                "sqref": str(dv.sqref) if dv.sqref else None,
            })
    
    # Named ranges from workbook
    if wb:
        try:
            if wb.defined_names:
                for name, dn in wb.defined_names.items():
                    try:
                        meta["named_ranges"].append({
                            "name": name,
                            "value": dn.value,
                        })
                    except Exception:
                        pass
        except Exception:
            pass
    
    return meta


def detect_input_cells(ws_val, ws_form, row_start, row_end, col_start, col_end):
    """
    Identify cells that contain values but NO formulas.
    These are 'manual input' cells — where a human enters data.
    
    Returns a list of input cell positions.
    """
    input_cells = []
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            val = ws_val.cell(row=r, column=c).value
            form = ws_form.cell(row=r, column=c).value
            
            if val is not None and isinstance(val, (int, float)):
                # Has a numeric value
                if form is None or not str(form).startswith('='):
                    # But no formula — this is a manually entered number
                    input_cells.append({
                        "cell": f"{get_column_letter(c)}{r}",
                        "row": r,
                        "column": c,
                        "value": val,
                    })
    
    return input_cells


def get_number_format_type(number_format):
    """
    Classify a number format string into a human-readable type.
    """
    if not number_format or number_format == "General":
        return "general"
    
    nf_lower = number_format.lower()
    
    if "%" in number_format:
        return "percentage"
    elif "$" in number_format or "currency" in nf_lower:
        return "currency"
    elif any(x in nf_lower for x in ["yy", "mm", "dd"]):
        return "date"
    elif any(x in nf_lower for x in ["hh", "ss"]):
        return "time"
    elif "#,##0" in number_format or "0.00" in number_format:
        return "number"
    elif "0" in number_format:
        return "number"
    
    return "custom"
