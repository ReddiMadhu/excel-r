import os
import argparse
import json
import glob
import traceback

import src.extractors.workbook_loader as workbook_loader
import src.extractors.sheet_classifier as sheet_classifier
import src.parsers.raw_data_parser as raw_data_parser
import src.parsers.pivot_parser as pivot_parser
import src.parsers.summary_table_detector as summary_table_detector
import src.core.json_builder as json_builder
import src.utils.validation as validation
import src.parsers.formula_parser as formula_parser
from src.utils.timing_log import PipelineTimer

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

def process_single_file(file_path, output_dir):
    """Process a single Excel workbook and output its JSON."""
    file_name = os.path.basename(file_path)
    print(f"Processing workbook: {file_name}")

    timer = PipelineTimer("parse_workbook", file_name=file_name)

    with timer.step("md5_hash"):
        file_hash = workbook_loader.compute_md5(file_path)

    with timer.step("load_workbook_values"):
        wb_val = workbook_loader.load_workbook_values(file_path)

    with timer.step("load_workbook_formulas"):
        wb_form = workbook_loader.load_workbook_formulas(file_path)

    with timer.step("classify_sheets"):
        sheet_types = sheet_classifier.classify_all_sheets(wb_val, wb_form=wb_form)

    # Identify summary and raw sheet names
    summary_sheet_name = None
    raw_sheet_name = None
    for name, s_type in sheet_types.items():
        if s_type == "summary_report":
            summary_sheet_name = name
        elif s_type == "raw_data":
            raw_sheet_name = name

    if not summary_sheet_name:
        timer.finish("PARSE_TOTAL_ERROR")
        raise ValueError(f"No summary/report sheet could be detected in {file_name}.")

    with timer.step("parse_column_headers"):
        from openpyxl.utils import get_column_letter
        raw_column_maps = {}
        for sheet_name in wb_val.sheetnames:
            ws = wb_val[sheet_name]
            if ws.max_row > 0 and ws.max_column > 0:
                header_row = 1
                try:
                    header_row = raw_data_parser.detect_header_row(ws)
                except Exception:
                    pass
                mapping = {}
                for col_idx in range(1, ws.max_column + 1):
                    val = ws.cell(row=header_row, column=col_idx).value
                    if val is not None:
                        col_letter = get_column_letter(col_idx)
                        mapping[col_letter] = str(val).strip()
                raw_column_maps[sheet_name] = mapping

    with timer.step("parse_pivot_metadata"):
        pivots_meta = []
        try:
            pivots_meta = pivot_parser.parse_pivot_metadata(file_path)
        except Exception as e:
            print(f"Warning extracting zip XML pivots for {file_name}: {e}")

    with timer.step("extract_summary_tables"):
        detected_tables_by_sheet = {}
        for name, s_type in sheet_types.items():
            if s_type == "summary_report":
                ws_v = wb_val[name]
                ws_f = wb_form[name]
                tables = summary_table_detector.extract_tables_from_sheet(
                    ws_v, ws_f, pivots_meta, wb_val
                )
                detected_tables_by_sheet[name] = tables

    detected_tables = []
    for tables in detected_tables_by_sheet.values():
        detected_tables.extend(tables)

    with timer.step("formulas_library_compile"):
        xl_model = None
        try:
            max_cells = int(os.getenv("FORMULAS_LIB_MAX_CELLS", "5000"))
            if max_cells > 0:
                xl_model = formula_parser.compile_workbook_scoped(
                    file_path,
                    sheet_types=sheet_types,
                    detected_tables=detected_tables,
                    max_cells=max_cells,
                )
        except Exception as e:
            print(f"Warning: formulas library budget compile failed for {file_name}: {e}")

    with timer.step("build_workbook_json"):
        output_json = json_builder.build_workbook_json(
            file_path,
            file_hash,
            sheet_types,
            wb_val,
            wb_form,
            raw_column_maps,
            pivots_meta,
            detected_tables,
            xl_model
        )

    with timer.step("validate_and_save_json"):
        warnings = validation.validate_extracted_json(output_json)
        output_json["extraction_warnings"] = warnings
        output_json["comparison_readiness"] = validation.compute_comparison_readiness(output_json)

        base_name = os.path.splitext(file_name)[0]
        out_path = os.path.join(output_dir, f"{base_name}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(output_json, f, indent=4)

    timer.finish("PARSE_TOTAL")
    print(f"Saved JSON: {out_path} (Warnings count: {len(warnings)})")
    return warnings

def generate_logical_steps_doc(output_dir):
    """Write logical_steps.md explaining how the code works."""
    logical_steps_content = """# Logical Steps for Excel-to-JSON Extractor

This document explains the extraction logic used by the Python Excel-to-JSON parser to read, scan, and convert reporting/summary sheets into structured JSON.

## 1. Sheet Classification
- **Summary Sheets**: Detected dynamically if the sheet name contains the word "summary" or "report" (case-insensitive).
- **Raw/Data Sheets**: Identified by checking for names like `Synthetic_Data`, `SQL_data`, or `Warehouse Data`. If none of these match, the sheet with the highest row count that contains a structured tabular header is classified as raw data.
- **Helper/Notes Sheets**: Identified as sheets named `Sheet1`, `helper`, or `notes` that have fewer than 100 rows.

## 2. Raw Data Mapping
- The header row is dynamically located on the raw data sheet by scanning the first 20 rows and identifying the first row containing multiple unique text values.
- A mapping is constructed between each column letter (e.g., `D`, `G`, `L`) and its header name (e.g., `Statutory Reserves - Total - General Account`).
- This mapping allows the formula parser to resolve column letters inside functions to their descriptive raw headers.

## 3. Pivot Table Processing
- Since Excel pivot table metadata is not fully exposed by openpyxl, the script opens the `.xlsx` file as a ZIP archive.
- It parses `xl/pivotTables/pivotTable1.xml` and `xl/pivotCache/pivotCacheDefinition1.xml` to extract:
  - Pivot table range (e.g., `A6:I67`).
  - Active Row fields, Column fields, and Page fields (filters).
  - Value/data fields and their aggregations (e.g., `SUM`, `COUNT`).
- A human-readable pivot formula is reconstructed for each value field (e.g., `SUM(Synthetic_Data[GA Stat Reserve]) GROUP BY Business Unit, ...`).
- Page fields (filters) are extracted from cells above the pivot range on the Summary sheet (e.g., `A1:B4` containing filter states) and stored under `filters` at the sheet level.

## 4. Advanced Table boundary Detection
Summary sheets contain plain ranges instead of formal Excel table objects. The parser scans the Summary sheet vertically and horizontally:
- **Vertical Partitioning**: Contiguous rows containing text or formulas separated by blank rows are identified as vertical blocks.
- **Horizontal Slicing**: Within each vertical block, columns that are completely blank in that row range are treated as horizontal separators. This splits rows into horizontal table blocks (e.g., columns `A:G` and columns `I:K` become separate tables).
- **Table Range Classification**:
  - The row range of each table block is partitioned into:
    - **Title Row**: Row at the top with exactly one non-empty string.
    - **Header Row**: Rows with text headers below the title.
    - **Data Rows**: Rows containing numeric values or formulas.
    - **Total/Check Rows**: Rows starting with "Total", "Grand Total", or "Check" are kept within their parent table boundaries.
- **Multi-Header Identification**:
  - If multiple header rows are detected above the data block, the text values are concatenated (e.g., `STAT Reserve` on row 36 and `Net Reserve` on row 37 are combined into `STAT Reserve Net Reserve`) to prevent duplicate header warnings.

## 5. Formula Parsing and Column Lineage
- Formulas are loaded using `data_only=False` to read raw formulas.
- We tokenize and parse `SUMIFS`, `COUNTIFS`, `SUM`, and arithmetic formulas.
- Cell and column letter references are mapped:
  - References like `SQL_data!D:D` are converted to `SQL_data[Statutory Reserves - Total - General Account]`.
  - Local cell references like `$A6` on the same row are replaced with `current Product Subtype`.
  - Local cell references to other rows/values are replaced with their evaluated cell value (e.g., `Summary Date ("12/31/2025")`).
- Calculated pivot values are marked as `type = pivot_value` and mapped to their source columns in the raw data sheet.
- Column definitions are automatically generated based on the formula patterns and lineage metadata.

## 6. Forward-Filling and JSON Assembly
- Parent label fields (such as `Business Unit` or `Statutory Exhibit section`) are forward-filled row-by-row inside `row_label_values` to provide local cell lineage.
- Original cell blanks are preserved in the row `values` dictionary.
- The raw data sheets are excluded from the output `sheets` array and kept only for internal lineage resolution, respecting the strict output scope rule.
"""
    out_path = os.path.join(output_dir, "logical_steps.md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(logical_steps_content)
    print(f"Saved logical steps document: {out_path}")

def main():
    parser = argparse.ArgumentParser(description="Excel-to-JSON Summary Extractor")
    parser.add_argument("--input_dir", default="data/input", help="Directory containing Excel files")
    parser.add_argument("--output_dir", default="data/output", help="Directory to save JSON output files")
    parser.add_argument("--file", help="Path to a single specific Excel file to process")
    args = parser.parse_args()
    
    input_dir = os.path.abspath(args.input_dir)
    output_dir = os.path.abspath(args.output_dir)
    
    os.makedirs(output_dir, exist_ok=True)
    
    if args.file:
        file_path = args.file
        if not os.path.exists(file_path):
            file_path = os.path.join(input_dir, args.file)
            if not os.path.exists(file_path):
                print(f"Error: File not found: {args.file} or {file_path}")
                return
        file_path = os.path.abspath(file_path)
        files_to_process = [file_path]
        print(f"Processing single file: {file_path}")
    else:
        print(f"Input directory: {input_dir}")
        print(f"Output directory: {output_dir}")
        
        # Look for .xlsx files in the input directory, ignoring temporary/autosave files
        files_to_process = [
            f for f in glob.glob(os.path.join(input_dir, "*.xlsx"))
            if not os.path.basename(f).startswith("~$")
        ]
        
    print(f"Found {len(files_to_process)} file(s) to process.")
    
    all_warnings = {}
    
    for file_path in files_to_process:

        fn = os.path.basename(file_path)
        try:
            warnings = process_single_file(file_path, output_dir)
            all_warnings[fn] = warnings
        except Exception as e:
            print(f"ERROR processing file {fn}: {e}")
            traceback.print_exc()
            all_warnings[fn] = [f"Critical error during processing: {str(e)}", traceback.format_exc()]
            
    # Generate validation report
    validation_report = validation.generate_validation_report(all_warnings)
    report_path = os.path.join(output_dir, "validation_report.json")
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(validation_report, f, indent=4)
        
    print(f"Saved validation report: {report_path}")
    
    # Generate logical steps document
    generate_logical_steps_doc(output_dir)
    
    print("Excel-to-JSON extraction completed successfully.")

if __name__ == "__main__":
    main()
