"""
Excel Review engine — per-workbook cell/formula inspection.

Runs after Discovery extraction. Independent of portfolio rationalization.
A single workbook MUST still produce Excel Review findings when warranted.
"""
from __future__ import annotations

import json
import logging
import os
from typing import Any, Dict, List, Optional, Set

from src.server.models.database import Database

logger = logging.getLogger(__name__)


def _load_defined_names(wb) -> Set[str]:
    names: Set[str] = set()
    try:
        if getattr(wb, "defined_names", None):
            for name in wb.defined_names:
                names.add(str(name))
    except Exception:
        pass
    return names


def _insert_finding(db: Database, workbook_id: int, finding: Dict[str, Any]) -> int:
    return db.insert("excel_review_findings", {
        "workbook_id": workbook_id,
        "finding_type": finding["finding_type"],
        "sheet": finding.get("sheet"),
        "cell": finding.get("cell"),
        "severity": finding.get("severity", "MEDIUM"),
        "confidence": finding.get("confidence", "medium"),
        "confidence_score": finding.get("confidence_score"),
        "actual": finding.get("actual"),
        "expected_pattern": finding.get("expected_pattern"),
        "evidence": json.dumps(finding.get("evidence") or []),
        "dependencies": json.dumps(finding.get("dependencies") or []),
        "dependents": json.dumps(finding.get("dependents") or []),
        "signals": json.dumps(finding.get("signals") or {}),
        "status": finding.get("status", "OPEN"),
    })


def _degraded_lineage_findings(db: Database, workbook_id: int) -> List[Dict[str, Any]]:
    rows = db.query(
        """
        SELECT column_name, table_name, formula, resolved_by
        FROM columns
        WHERE workbook_id = ?
          AND column_type IN ('formula_based', 'pivot_value', 'total')
          AND (resolved_by IN ('degraded', 'unsupported')
               OR formula_lineage IS NULL OR formula_lineage = '' OR formula_lineage = '{}')
        LIMIT 30
        """,
        (workbook_id,),
    )
    out: List[Dict[str, Any]] = []
    for col in rows:
        out.append({
            "finding_type": "DEGRADED_LINEAGE",
            "sheet": col.get("table_name"),
            "cell": None,
            "severity": "LOW",
            "confidence": "medium",
            "confidence_score": 0.6,
            "actual": col.get("formula"),
            "expected_pattern": None,
            "evidence": [
                f"Column '{col.get('column_name')}' in '{col.get('table_name')}' "
                f"has degraded or missing lineage (resolved_by={col.get('resolved_by')}).",
            ],
            "dependencies": [],
            "dependents": [],
            "signals": {
                "column": col.get("column_name"),
                "resolved_by": col.get("resolved_by"),
            },
        })
    return out


def run_excel_review(
    db: Database,
    workbook_id: int,
    file_path: Optional[str] = None,
    *,
    max_cells_per_sheet: int = 100_000,
) -> Dict[str, Any]:
    """
    Analyze one workbook and write excel_review_findings.

    Returns summary dict with status and counts.
    """
    from src.review.region_detector import detect_regions
    from src.review.override_detector import detect_all_overrides
    from src.review.ref_auditor import audit_all_cells

    summary: Dict[str, Any] = {
        "status": "completed",
        "workbook_id": workbook_id,
        "findings": 0,
        "by_type": {},
        "warnings": [],
        "phase_errors": [],
    }

    if not file_path:
        wb_row = db.query_one(
            "SELECT source_file, name FROM workbooks WHERE id = ?", (workbook_id,)
        )
        if not wb_row:
            summary["status"] = "failed"
            summary["phase_errors"].append({"phase": "load", "error": "workbook not found"})
            return summary
        file_path = wb_row.get("source_file")

    if not file_path or not os.path.exists(file_path):
        summary["status"] = "failed"
        summary["phase_errors"].append({
            "phase": "load",
            "error": f"source file missing: {file_path}",
        })
        return summary

    db.execute("DELETE FROM excel_review_findings WHERE workbook_id = ?", (workbook_id,))

    try:
        import openpyxl
        wb_form = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        try:
            wb_val = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as e:
            wb_val = None
            summary["warnings"].append(f"data_only load failed: {e}")
    except Exception as e:
        logger.exception("Failed to open workbook for Excel Review: %s", e)
        summary["status"] = "failed"
        summary["phase_errors"].append({"phase": "open", "error": str(e)})
        return summary

    defined_names = _load_defined_names(wb_form)
    all_findings: List[Dict[str, Any]] = []
    sheets_scanned = 0
    unsupported_workbook = False

    try:
        raw_keywords = ["data", "raw", "extract", "dump", "source", "query", "sql", "synthetic"]
        for ws_form in wb_form.worksheets:
            title_lower = ws_form.title.lower()
            # Skip pure raw data sheets in cell-level review (they contain no report formulas to audit)
            if any(kw in title_lower for kw in raw_keywords) and ws_form.max_row and ws_form.max_row > 100:
                logger.info("Excel Review: skipping raw data sheet '%s' (%d rows)", ws_form.title, ws_form.max_row)
                continue

            sheets_scanned += 1
            ws_val = None
            if wb_val is not None:
                try:
                    ws_val = wb_val[ws_form.title]
                except KeyError:
                    ws_val = None
            try:
                cells, regions = detect_regions(
                    ws_form, ws_val, sheet_name=ws_form.title,
                    max_cells=max_cells_per_sheet,
                )
            except Exception as e:
                logger.exception("Region detection failed on sheet %s", ws_form.title)
                summary["warnings"].append(f"region_detection:{ws_form.title}: {e}")
                continue

            try:
                all_findings.extend(detect_all_overrides(regions))
            except Exception as e:
                summary["warnings"].append(f"override_detection:{ws_form.title}: {e}")

            try:
                ref_findings = audit_all_cells(cells, defined_names=defined_names)
                all_findings.extend(ref_findings)
                if any(f["finding_type"] == "UNSUPPORTED_FEATURE" for f in ref_findings):
                    unsupported_workbook = True
            except Exception as e:
                summary["warnings"].append(f"ref_audit:{ws_form.title}: {e}")
    finally:
        try:
            wb_form.close()
        except Exception:
            pass
        if wb_val is not None:
            try:
                wb_val.close()
            except Exception:
                pass

    try:
        all_findings.extend(_degraded_lineage_findings(db, workbook_id))
    except Exception as e:
        summary["warnings"].append(f"degraded_lineage_query: {e}")

    # Dedupe
    final_findings: List[Dict[str, Any]] = []
    seen = set()
    for f in all_findings:
        signals = f.get("signals") if isinstance(f.get("signals"), dict) else {}
        key = (
            f["finding_type"],
            f.get("sheet"),
            f.get("cell"),
            f.get("actual"),
            signals.get("column"),
        )
        if key in seen:
            continue
        seen.add(key)
        final_findings.append(f)

    for finding in final_findings:
        try:
            _insert_finding(db, workbook_id, finding)
        except Exception as e:
            summary["phase_errors"].append({"phase": "write", "error": str(e)})
            summary["status"] = "partial"

    by_type: Dict[str, int] = {}
    for f in final_findings:
        by_type[f["finding_type"]] = by_type.get(f["finding_type"], 0) + 1

    summary["findings"] = len(final_findings)
    summary["by_type"] = by_type
    summary["sheets_scanned"] = sheets_scanned
    summary["unsupported_features_present"] = unsupported_workbook

    if summary["phase_errors"] and summary["findings"] == 0:
        summary["status"] = "failed"
    elif summary["phase_errors"] or summary["warnings"]:
        if summary["status"] != "partial":
            summary["status"] = "completed_with_warnings"

    logger.info(
        "Excel Review workbook_id=%s status=%s findings=%d",
        workbook_id, summary["status"], summary["findings"],
    )
    return summary


def run_excel_review_for_scan(
    db: Database,
    workbook_ids: Optional[List[int]] = None,
) -> Dict[str, Any]:
    """Run Excel Review for all (or selected) workbooks."""
    if workbook_ids:
        placeholders = ",".join("?" * len(workbook_ids))
        rows = db.query(
            f"SELECT id, source_file FROM workbooks WHERE id IN ({placeholders})",
            tuple(workbook_ids),
        )
    else:
        rows = db.query("SELECT id, source_file FROM workbooks")

    results = []
    for row in rows:
        results.append(run_excel_review(db, row["id"], row.get("source_file")))

    total = sum(r.get("findings", 0) for r in results)
    statuses = {r.get("status") for r in results}
    if "failed" in statuses and total == 0:
        status = "failed"
    elif "failed" in statuses or "partial" in statuses or "completed_with_warnings" in statuses:
        status = "completed_with_warnings" if "failed" not in statuses else "partial"
    else:
        status = "completed"
    return {
        "status": status,
        "workbooks": len(results),
        "findings": total,
        "results": results,
    }
