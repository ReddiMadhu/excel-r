"""
Formula region detection — O(n) grouping by contiguous formula blocks.

Understands:
  - contiguous formula rows/columns
  - blank separator rows
  - mixed formula/constant rows (override candidates live here)
Does NOT flag every header/subtotal boundary as an error.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.utils import get_column_letter

from src.review.pattern_normalizer import (
    is_constant_nonblank,
    is_formula,
    structural_pattern_key,
)


@dataclass
class CellInfo:
    sheet: str
    row: int
    col: int  # 1-based
    address: str
    formula: Optional[str]
    value: Any
    pattern_key: str = ""
    unsupported: List[str] = field(default_factory=list)


@dataclass
class FormulaRegion:
    sheet: str
    # Axis of the region: "row" means a contiguous run of cells in one row;
    # "col" means a contiguous run down one column (fill-down).
    axis: str
    fixed_index: int  # row number if axis=row, col number if axis=col
    start: int
    end: int
    cells: List[CellInfo] = field(default_factory=list)
    dominant_pattern: Optional[str] = None
    pattern_counts: Dict[str, int] = field(default_factory=dict)


def _scan_sheet_cells(ws_form, ws_val, sheet_name: str, max_cells: int = 50_000) -> List[CellInfo]:
    """Collect non-empty cells with formulas or values (streamed & capped for performance)."""
    cells: List[CellInfo] = []
    count = 0
    
    # Use streaming iter_rows for high-performance sequential scanning
    form_rows = ws_form.iter_rows(values_only=False)
    val_rows = ws_val.iter_rows(values_only=True) if ws_val is not None else None
    
    for r_idx, form_row in enumerate(form_rows, start=1):
        val_row = next(val_rows, None) if val_rows is not None else None
        
        for c_idx, form_cell in enumerate(form_row, start=1):
            raw = getattr(form_cell, "value", None)
            if raw is None or raw == "":
                continue
            
            val = None
            if val_row and (c_idx - 1) < len(val_row):
                val = val_row[c_idx - 1]
                
            addr = f"{get_column_letter(c_idx)}{r_idx}"
            formula = str(raw) if is_formula(raw) else None
            pattern = ""
            unsupported: List[str] = []
            if formula:
                pattern, unsupported = structural_pattern_key(formula, r_idx, c_idx)
            cells.append(CellInfo(
                sheet=sheet_name,
                row=r_idx,
                col=c_idx,
                address=addr,
                formula=formula,
                value=val if formula is None else (val if val is not None else raw),
                pattern_key=pattern,
                unsupported=unsupported,
            ))
            count += 1
            if count >= max_cells:
                return cells
    return cells


def _contiguous_runs(
    items: List[CellInfo],
    axis: str,
) -> List[List[CellInfo]]:
    """
    Group cells into contiguous runs along an axis.
    axis='row': group by row, contiguous columns.
    axis='col': group by column, contiguous rows.
    """
    buckets: Dict[int, List[CellInfo]] = defaultdict(list)
    for cell in items:
        key = cell.row if axis == "row" else cell.col
        buckets[key].append(cell)

    runs: List[List[CellInfo]] = []
    for fixed, group in buckets.items():
        group.sort(key=lambda x: x.col if axis == "row" else x.row)
        current: List[CellInfo] = []
        prev = None
        for cell in group:
            idx = cell.col if axis == "row" else cell.row
            if prev is None or idx == prev + 1:
                current.append(cell)
            else:
                if len(current) >= 2:
                    runs.append(current)
                current = [cell]
            prev = idx
        if len(current) >= 2:
            runs.append(current)
    return runs


def _region_from_run(run: List[CellInfo], axis: str) -> FormulaRegion:
    fixed = run[0].row if axis == "row" else run[0].col
    start = run[0].col if axis == "row" else run[0].row
    end = run[-1].col if axis == "row" else run[-1].row
    pattern_counts: Dict[str, int] = defaultdict(int)
    for c in run:
        if c.pattern_key:
            pattern_counts[c.pattern_key] += 1
        elif c.formula is None and is_constant_nonblank(c.value):
            pattern_counts["__CONSTANT__"] += 1
    dominant = None
    if pattern_counts:
        # Prefer formula patterns over constant marker
        formula_patterns = {k: v for k, v in pattern_counts.items() if k != "__CONSTANT__"}
        if formula_patterns:
            dominant = max(formula_patterns, key=formula_patterns.get)
        else:
            dominant = "__CONSTANT__"
    return FormulaRegion(
        sheet=run[0].sheet,
        axis=axis,
        fixed_index=fixed,
        start=start,
        end=end,
        cells=run,
        dominant_pattern=dominant,
        pattern_counts=dict(pattern_counts),
    )


def detect_regions(
    ws_form,
    ws_val=None,
    sheet_name: str = "",
    max_cells: int = 200_000,
) -> Tuple[List[CellInfo], List[FormulaRegion]]:
    """
    Scan a worksheet and return all interesting cells + formula regions.

    Regions are built from both row-wise and column-wise contiguous runs
    that contain at least one formula (pure constant blocks are skipped).
    """
    name = sheet_name or getattr(ws_form, "title", "Sheet")
    cells = _scan_sheet_cells(ws_form, ws_val, name, max_cells=max_cells)

    # Include blanks between formulas? We only keep non-empty cells.
    # For override detection we also need constants that sit inside formula spans —
    # those are already in `cells` if non-blank.

    formula_cells = [c for c in cells if c.formula]
    # Also keep constants for region adjacency analysis
    interesting = list(cells)

    regions: List[FormulaRegion] = []
    for axis in ("row", "col"):
        for run in _contiguous_runs(interesting, axis):
            if any(c.formula for c in run):
                regions.append(_region_from_run(run, axis))

    return cells, regions
