"""
JSON Builder — Build the structured JSON output for a workbook.

Enhanced with:
- Cell formatting metadata (bold, borders, indent, number format)
- Row hierarchy from indentation
- Input cell detection (manual vs formula-driven)
- Print area and named ranges in metadata
- formulas library dependency resolution
- resolved_by tracking for each column
"""
import datetime
import openpyxl
import os
import time
import re
import json
from openpyxl.utils import get_column_letter
import formula_parser
import pivot_parser
import formatting_extractor
import workbook_loader
import structural_summary
import summary_table_detector
import llm_client


HAS_GENAI_LIB = None
HAS_OPENAI_LIB = None
HAS_ANTHROPIC_LIB = None


def normalize_key(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower()) if s else ""


def analyze_workbook_semantics_llm(skeleton):
    """
    Generate workbook, table, and column definitions using the resilient llm_client.
    Sends the entire skeleton JSON to the LLM in a single request.
    """
    import json
    
    llm = llm_client.get_resilient_llm(temperature=0.3, json_mode=True)
    if llm is None:
        return None
        
    skeleton_str = json.dumps(skeleton, indent=2)
    
    prompt = f"""You are an expert actuarial and financial business analyst.
Analyze the following Excel workbook structural skeleton and answer all semantic questions in a single structured JSON response.

Workbook Skeleton:
{skeleton_str}

Please generate a JSON object with the following schema:
{{
  "workbook_purpose": "A concise (2-3 sentences) description of the business purpose and goals of this workbook.",
  "process_flow": {{
    "primary_inputs": ["SheetName or TableName that acts as raw inputs"],
    "intermediate_calculations": ["TableName or SheetName representing calculations"],
    "final_outputs": ["TableName or SheetName representing the final deliverables or prints"],
    "vulnerability_rating": "low, medium, or high (based on presence of hardcoded values, complex loops, or lack of checks)"
  }},
  "tables": [
    {{
      "table_name": "Name of the table",
      "table_definition": "A concise (1-2 sentences) business definition of what this table represents.",
      "boundary_correction": {{
        "is_correct": true,
        "suggested_header_rows": null,
        "suggested_total_rows": null
      }},
      "inter_table_relationships": [
        "Description of how this table validates or references other tables (e.g., check cells)"
      ],
      "columns": [
        {{
          "column_name": "Name of the column",
          "column_definition": "A concise (one sentence) explanation of what this column represents and how it is derived from a business perspective."
        }}
      ]
    }}
  ]
}}

Your response must be a single raw JSON object matching the schema above. Do not include any introductory remarks, markdown wraps, or explanations outside the JSON. Start directly with the opening curly brace '{{'."""

    try:
        response = llm.invoke(prompt)
        response_text = llm_client.stringify_chat_content(response.content).strip()
        
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        elif response_text.startswith("```"):
            response_text = response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        response_text = response_text.strip()
        
        return json.loads(response_text)
    except Exception as e:
        print(f"Warning: Failed calling resilient LLM or parsing response: {e}")
        return None


def normalize_pivot_metadata_only(sheet_json):
    """
    Post-processing function to normalize pivot metadata:
    1. Collect existing pivot metadata from sheet-level or columns/tables.
    2. Deduplicate pivot tables by pivot_table_name.
    3. Ensure sheet_json["pivot_tables"] contains the unique list of pivot tables.
    4. Replace duplicated column-level full pivot objects with lightweight references.
    """
    pivot_tables_by_name = {}
    
    existing_pivots = sheet_json.get("pivot_tables", [])
    if existing_pivots is None:
        existing_pivots = []
    for pt in existing_pivots:
        name = pt.get("pivot_table_name") or pt.get("name")
        if name:
            pivot_tables_by_name[name] = pt

    tables = sheet_json.get("tables", [])
    for table in tables:
        t_name = table.get("table_name", "")
        t_type = table.get("table_type", "")
        
        table_pivot_meta = None
        keys_to_remove = []
        for k, v in list(table.items()):
            if "pivot" in k.lower() and isinstance(v, dict) and ("pivot_table_name" in v or "value_fields" in v):
                table_pivot_meta = v
                keys_to_remove.append(k)
        
        for k in keys_to_remove:
            del table[k]
            
        if table_pivot_meta:
            pt_name = table_pivot_meta.get("pivot_table_name") or table_pivot_meta.get("name") or t_name
            if pt_name not in pivot_tables_by_name:
                if "pivot_table_name" not in table_pivot_meta and "name" in table_pivot_meta:
                    table_pivot_meta["pivot_table_name"] = table_pivot_meta.pop("name")
                pivot_tables_by_name[pt_name] = table_pivot_meta

        for col in table.get("columns", []):
            col_type = col.get("type", "")
            
            col_pivot_meta = None
            col_keys_to_remove = []
            for k, v in list(col.items()):
                if "pivot" in k.lower() and isinstance(v, dict) and ("pivot_table_name" in v or "value_fields" in v):
                    col_pivot_meta = v
                    col_keys_to_remove.append(k)
                    
            for k in col_keys_to_remove:
                del col[k]
                
            pt_name = None
            if col_pivot_meta:
                pt_name = col_pivot_meta.get("pivot_table_name") or col_pivot_meta.get("name") or t_name
                if pt_name not in pivot_tables_by_name:
                    if "pivot_table_name" not in col_pivot_meta and "name" in col_pivot_meta:
                        col_pivot_meta["pivot_table_name"] = col_pivot_meta.pop("name")
                    pivot_tables_by_name[pt_name] = col_pivot_meta
            elif col_type == "pivot_value" or t_type == "pivot_table":
                pt_name = t_name
                
            if pt_name:
                col["related_pivot_table_names"] = [pt_name]

    sheet_json["pivot_tables"] = list(pivot_tables_by_name.values())
    
    if "pivot_tables" not in sheet_json or sheet_json["pivot_tables"] is None:
        sheet_json["pivot_tables"] = []


def build_workbook_json(file_name, file_hash, sheet_classifications, wb_val, wb_form, 
                        raw_column_maps, pivot_tables_meta, detected_tables, xl_model=None):
    """
    Build the structured JSON for a workbook following the expected schema.
    
    Enhanced with:
    - xl_model: compiled formulas library model for dependency resolution
    - Formatting metadata per column/row
    - Hierarchy information from indentation
    - Input cell detection
    - Print area and named range metadata
    """
    sheets_json = []
    
    summary_sheet_name = ""
    raw_data_sheet_name = ""
    for s_name, s_type in sheet_classifications.items():
        if s_type == "summary_report":
            summary_sheet_name = s_name
        elif s_type == "raw_data":
            raw_data_sheet_name = s_name

    # Extract workbook-level metadata
    wb_meta = workbook_loader.extract_workbook_metadata(wb_val, file_path=file_name)

    # Build skeleton and analyze semantics via LLM
    skeleton = structural_summary.generate_workbook_skeleton(
        file_name, sheet_classifications, wb_val, wb_form, raw_column_maps, detected_tables
    )
    semantics = analyze_workbook_semantics_llm(skeleton)
    
    workbook_purpose = ""
    process_flow = {}
    table_semantics = {}
    
    if semantics:
        workbook_purpose = semantics.get("workbook_purpose", "")
        process_flow = semantics.get("process_flow", {})
        for tbl_sem in semantics.get("tables", []):
            t_name_sem = tbl_sem.get("table_name", "")
            t_key = normalize_key(t_name_sem)
            col_dict = {}
            for col_sem in tbl_sem.get("columns", []):
                c_name_sem = col_sem.get("column_name", "")
                c_key = normalize_key(c_name_sem)
                col_dict[c_key] = col_sem.get("column_definition", "")
            table_semantics[t_key] = {
                "definition": tbl_sem.get("table_definition", ""),
                "boundary_correction": tbl_sem.get("boundary_correction", {}),
                "inter_table_relationships": tbl_sem.get("inter_table_relationships", []),
                "columns": col_dict
            }

    for s_name, s_type in sheet_classifications.items():
        ws_val = wb_val[s_name]
        ws_form = wb_form[s_name]
        
        # Calculate metadata
        row_count = ws_val.max_row or 0
        col_count = ws_val.max_column or 0
        
        formula_count = 0
        non_empty_count = 0
        
        # Fast iteration using iter_rows(values_only=True)
        for row in ws_val.iter_rows(values_only=True):
            for val_v in row:
                if val_v is not None:
                    non_empty_count += 1
                    
        for row in ws_form.iter_rows(values_only=True):
            for val_f in row:
                if val_f is not None and str(val_f).startswith('='):
                    formula_count += 1

        # Extract sheet-level metadata (print area, hidden rows/cols)
        sheet_meta = workbook_loader.extract_sheet_metadata(ws_val)
        
        # Cache column values/formulas to prevent slow cell-by-cell access on large sheets
        col_values_cache = {}
        col_formulas_cache = {}
        if row_count > 1000:
            all_rows_val = list(ws_val.iter_rows(values_only=True))
            all_rows_form = list(ws_form.iter_rows(values_only=True))
            for c_idx in range(1, col_count + 1):
                col_values_cache[c_idx] = [
                    row[c_idx - 1] if c_idx - 1 < len(row) else None
                    for row in all_rows_val
                ]
                col_formulas_cache[c_idx] = [
                    row[c_idx - 1] if c_idx - 1 < len(row) else None
                    for row in all_rows_form
                ]

        # Extract filters
        filters_list = []
        if s_type == "summary_report":
            pivot_filters = {}
            for pt in pivot_tables_meta:
                reconstructed = pivot_parser.reconstruct_pivot_formulas(pt, ws_val)
                if reconstructed:
                    first_measure = list(reconstructed.values())[0]
                    pivot_filters = first_measure.get("pivot_filters", {})
                    break
                    
            for k, v in pivot_filters.items():
                filters_list.append({
                    "filter_name": k,
                    "filter_value": v
                })
                
            if not filters_list:
                for r in range(1, 5):
                    cell_a = ws_val.cell(row=r, column=1).value
                    cell_b = ws_val.cell(row=r, column=2).value
                    if cell_a is not None and cell_b is not None:
                        if isinstance(cell_a, str) and not cell_a.replace(' ', '').isalnum():
                            pass
                        elif isinstance(cell_a, str):
                            filters_list.append({
                                "filter_name": str(cell_a).strip(),
                                "filter_value": str(cell_b).strip()
                            })

        # Extract columns list for backward compatibility in raw/helper metadata
        columns_list = []
        if s_type != "summary_report":
            if s_type == "raw_data":
                if s_name in raw_column_maps:
                    columns_list = list(raw_column_maps[s_name].values())
                else:
                    for c_idx in range(1, col_count + 1):
                        val_v = ws_val.cell(row=1, column=c_idx).value
                        if val_v is not None:
                            columns_list.append(str(val_v).strip())
            elif s_type == "helper":
                for c_idx in range(1, min(col_count + 1, 30)):
                    val_v = ws_val.cell(row=1, column=c_idx).value
                    if val_v is not None:
                        columns_list.append(str(val_v).strip())

        # Determine tables for this sheet
        sheet_tables = []
        if s_type == "summary_report":
            sheet_tables = [t for t in detected_tables if t.get("table_range")]
        else:
            sheet_tables = summary_table_detector.extract_tables_from_sheet(
                ws_val, ws_form, pivot_tables_meta=None, wb=wb_val
            )
            # Default to treating raw data/helper sheets as a single table
            if not sheet_tables and row_count > 0 and col_count > 0:
                header_row = 1
                try:
                    import raw_data_parser
                    header_row = raw_data_parser.detect_header_row(ws_val)
                except Exception:
                    pass
                
                headers = []
                for c_idx in range(1, col_count + 1):
                    val_v = ws_val.cell(row=header_row, column=c_idx).value
                    headers.append(str(val_v).strip() if val_v is not None else f"Column_{get_column_letter(c_idx)}")
                
                sheet_tables = [{
                    "table_name": s_name,
                    "table_type": "standard_table",
                    "section_title": s_name,
                    "table_range": f"A1:{get_column_letter(col_count)}{row_count}",
                    "col_start": 1,
                    "col_end": col_count,
                    "row_start": 1,
                    "row_end": row_count,
                    "headers": headers,
                    "row_classification": {
                        "title_rows": list(range(1, header_row)) if header_row > 1 else [],
                        "header_rows": [header_row] if row_count >= header_row else [1],
                        "data_rows": list(range(header_row + 1, row_count + 1)) if row_count > header_row else [],
                        "total_rows": [],
                        "check_rows": []
                    }
                }]
                
        # Process tables
        tables_json = []
        for t_idx, tbl in enumerate(sheet_tables, 1):
            col_start = tbl["col_start"]
            col_end = tbl["col_end"]
            row_start = tbl["row_start"]
            row_end = tbl["row_end"]
            row_classification = tbl["row_classification"]
            
            t_name = tbl["table_name"]
            
            # Apply boundary correction if recommended by the LLM
            t_sem = table_semantics.get(normalize_key(t_name), {})
            boundary_corr = t_sem.get("boundary_correction", {})
            if boundary_corr and not boundary_corr.get("is_correct", True):
                suggested_headers = boundary_corr.get("suggested_header_rows")
                suggested_totals = boundary_corr.get("suggested_total_rows")
                if suggested_headers:
                    row_classification["header_rows"] = [int(r) for r in suggested_headers if str(r).isdigit()]
                if suggested_totals:
                    row_classification["total_rows"] = [int(r) for r in suggested_totals if str(r).isdigit()]
                
                t_title_rows = row_classification["title_rows"]
                t_header_rows = row_classification["header_rows"]
                t_total_rows = row_classification["total_rows"]
                t_check_rows = row_classification["check_rows"]
                
                all_possible = list(range(row_start, row_end + 1))
                data_rows = [r for r in all_possible if r not in t_title_rows and r not in t_header_rows and r not in t_total_rows and r not in t_check_rows]
                row_classification["data_rows"] = data_rows
                
                headers = summary_table_detector.build_disambiguated_headers(ws_val, t_header_rows, col_start, col_end)
                tbl["headers"] = headers

            t_type = tbl["table_type"]
            section_title = tbl.get("section_title", "")
            
            title_rows = row_classification["title_rows"]
            header_rows = row_classification["header_rows"]
            data_rows = row_classification["data_rows"]
            total_rows = row_classification["total_rows"]
            check_rows = row_classification["check_rows"]
            
            # Map columns to names
            headers = []
            if "headers" in tbl:
                headers = tbl["headers"]
            else:
                headers = [ws_val.cell(row=header_rows[-1] if header_rows else row_start, column=c).value for c in range(col_start, col_end + 1)]
                headers = [str(h).strip() if h is not None else f"Column_{get_column_letter(c)}" for c, h in zip(range(col_start, col_end + 1), headers)]
                
            # Disambiguate duplicate headers
            unique_headers = []
            seen_headers = {}
            for h in headers:
                if h in seen_headers:
                    seen_headers[h] += 1
                    unique_headers.append(f"{h}_{seen_headers[h]}")
                else:
                    seen_headers[h] = 0
                    unique_headers.append(h)
            headers = unique_headers

            # Identify row header columns
            row_header_cols = []
            if t_type == "pivot_table" and "pivot_meta" in tbl:
                row_header_cols = tbl["pivot_meta"].get("row_fields", [])
            else:
                val_start_col = col_start
                for c in range(col_start, col_end + 1):
                    col_has_values = False
                    for r in data_rows:
                        cell_val = ws_val.cell(row=r, column=c).value
                        cell_formula = ws_form.cell(row=r, column=c).value
                        if formula_parser.is_value_cell(cell_val, cell_formula):
                            col_has_values = True
                            break
                    if col_has_values:
                        val_start_col = c
                        break
                        
                for c in range(col_start, val_start_col):
                    hdr_name = headers[c - col_start]
                    hdr_lower = hdr_name.lower()
                    non_label_keywords = ["reserve", "yrt", "face", "count", "cash value", "amount", "flexible", "component", "total", "check"]
                    if not any(kw in hdr_lower for kw in non_label_keywords):
                        row_header_cols.append(hdr_name)
                
            # Table column index mapping
            table_col_mapping = {}
            for c in range(col_start, col_end + 1):
                table_col_mapping[c] = headers[c - col_start]
                
            # Populate columns JSON
            columns_json = []
            
            # Reconstruct pivot measures
            pivot_measures = {}
            if t_type == "pivot_table" and "pivot_meta" in tbl:
                pivot_measures = pivot_parser.reconstruct_pivot_formulas(tbl["pivot_meta"], ws_val)
                
            for c_idx, col_name in enumerate(headers, col_start):
                col_letter = get_column_letter(c_idx)
                
                # Get formulas from data rows
                if row_count > 1000:
                    col_formulas = [col_formulas_cache[c_idx][r - 1] for r in data_rows]
                else:
                    col_formulas = [ws_form.cell(row=r, column=c_idx).value for r in data_rows]
                formula_pattern_inferred = formula_parser.infer_formula_pattern_for_column(col_formulas)
                
                first_formula = next((f for f in col_formulas if f and str(f).startswith('=')), None)
                first_formula_row = next((r for r, f in zip(data_rows, col_formulas) if f and str(f).startswith('=')), None)
                
                # Try formulas library first, then fall back to custom parser
                parsed_f = {
                    "type": "raw",
                    "formula_pattern": "",
                    "data_source_sheet": "",
                    "data_source_columns": [],
                    "formula_source_details": [],
                    "formula_count": 0,
                    "resolved_by": "none",
                }
                
                if first_formula and first_formula_row:
                    cell_ref = f"{col_letter}{first_formula_row}"
                    
                    if xl_model is not None:
                        # Try formulas library first
                        parsed_f = formula_parser.parse_formula_with_library(
                            xl_model, s_name, cell_ref, first_formula,
                            first_formula_row, ws_val, raw_column_maps,
                            table_col_mapping, t_name, ws_form,
                            detected_tables=sheet_tables
                        )
                    else:
                        # Direct custom parsing
                        parsed_f = formula_parser.parse_formula(
                            first_formula, first_formula_row, ws_val,
                            raw_column_maps, table_col_mapping, t_name,
                            ws_form,
                            detected_tables=sheet_tables
                        )
                    
                # Samples & Data Type detection (limit scan for efficiency)
                sample_vals = []
                non_empty_for_dtype = []
                scan_limit = 1000
                checked_count = 0
                for r in data_rows:
                    if row_count > 1000:
                        v = col_values_cache[c_idx][r - 1]
                    else:
                        v = ws_val.cell(row=r, column=c_idx).value
                    if v is not None:
                        if len(sample_vals) < 20:
                            if isinstance(v, (datetime.datetime, datetime.date)):
                                sample_vals.append(v.isoformat())
                            else:
                                sample_vals.append(v)
                        if len(non_empty_for_dtype) < 100:
                            non_empty_for_dtype.append(v)
                        checked_count += 1
                        if checked_count >= scan_limit:
                            break
                            
                dtype = "string"
                if non_empty_for_dtype:
                    dtype = "float" if any(isinstance(x, float) for x in non_empty_for_dtype) else "integer" if any(isinstance(x, int) for x in non_empty_for_dtype) else "string"
                    
                # Determine column type
                col_type = "raw"
                if col_name in row_header_cols:
                    col_type = "label"
                elif t_type == "pivot_table" and col_name in pivot_measures:
                    col_type = "pivot_value"
                elif parsed_f["type"] == "formula_based":
                    col_type = "formula_based"
                elif parsed_f["type"] == "total":
                    col_type = "total"
                elif parsed_f["type"] == "check":
                    col_type = "check"
                    
                ds_sheet = parsed_f["data_source_sheet"]
                ds_cols = parsed_f["data_source_columns"]
                fs_details = parsed_f["formula_source_details"]
                f_pattern = parsed_f["formula_pattern"]
                f_count = parsed_f["formula_count"]
                resolved_by = parsed_f.get("resolved_by", "none")
                
                # Override with pivot details
                if col_type == "pivot_value" and col_name in pivot_measures:
                    p_details = pivot_measures[col_name]
                    ds_sheet = p_details["data_source_sheet"]
                    ds_cols = p_details["data_source_columns"]
                    f_pattern = p_details["formula_pattern"]
                    fs_details = []
                    for g_field in p_details["group_by_fields"]:
                        fs_details.append({"column_name": g_field, "role": "criteria_range"})
                    fs_details.append({"column_name": p_details["source_column"], "role": "sum_range" if p_details["aggregation"] == "SUM" else "count_range"})
                    resolved_by = "pivot_xml"
                    
                # Generate definitions
                definition = ""
                if col_type == "label":
                    definition = f"Row hierarchy label: {col_name}"
                elif col_type == "pivot_value" and col_name in pivot_measures:
                    p_details = pivot_measures[col_name]
                    definition = f"{p_details['aggregation']} of {p_details['source_column']} grouped by {', '.join(p_details['group_by_fields'])}"
                elif col_type == "formula_based":
                    if ds_sheet and ds_cols:
                        definition = f"Sum of {', '.join(ds_cols)} from {ds_sheet} grouped by {', '.join(row_header_cols)}"
                    else:
                        definition = f"Derived formula column: {f_pattern}"
                elif col_type == "total":
                    definition = f"Total sum of values for {col_name}"
                elif col_type == "check":
                    definition = f"Check validation variance for {col_name}"
                else:
                    definition = f"Raw summary value for {col_name}"
                    
                # Extract formatting for the header cell
                header_row_idx = header_rows[-1] if header_rows else row_start
                header_fmt = formatting_extractor.extract_cell_formatting(
                    ws_val.cell(row=header_row_idx, column=c_idx)
                )
                
                # Detect number format from data rows
                number_format = "General"
                for r in data_rows[:3]:
                    nf = ws_val.cell(row=r, column=c_idx).number_format
                    if nf and nf != "General":
                        number_format = nf
                        break
                    
                col_info = {
                    "column_name": col_name,
                    "excel_column": col_letter,
                    "table_name": t_name,
                    "data_type": dtype,
                    "type": col_type,
                    "data_source_sheet": ds_sheet,
                    "data_source_columns": ds_cols,
                    "formula_source_column_count": len(ds_cols),
                    "formula_source_details": fs_details,
                    "formula": first_formula or "",
                    "formula_count": f_count,
                    "formula_pattern": f_pattern,
                    "formula_applies_to": f"{col_letter}{data_rows[0]}:{col_letter}{data_rows[-1]}" if data_rows else "",
                    "resolved_by": resolved_by,
                    "number_format": number_format,
                    "number_format_type": formatting_extractor.get_number_format_type(number_format),
                    "sample_values": sample_vals[:20],
                }
                
                # Fetch LLM business definition from the whole-workbook semantic lookup
                col_sem = t_sem.get("columns", {}).get(normalize_key(col_name), "")
                if col_sem:
                    col_info["definition"] = col_sem
                else:
                    # Fallback to local heuristic definitions
                    if col_type == "label":
                        col_info["definition"] = f"Row hierarchy label: {col_name}"
                    elif col_type == "pivot_value" and col_name in pivot_measures:
                        p_details = pivot_measures[col_name]
                        col_info["definition"] = f"{p_details['aggregation']} of {p_details['source_column']} grouped by {', '.join(p_details['group_by_fields'])}"
                    elif col_type == "formula_based":
                        if ds_sheet and ds_cols:
                            col_info["definition"] = f"Sum of {', '.join(ds_cols)} from {ds_sheet} grouped by {', '.join(row_header_cols)}"
                        else:
                            col_info["definition"] = f"Derived formula column: {f_pattern}"
                    elif col_type == "total":
                        col_info["definition"] = f"Total sum of values for {col_name}"
                    elif col_type == "check":
                        col_info["definition"] = f"Check validation variance for {col_name}"
                    else:
                        col_info["definition"] = f"Raw summary value for {col_name}"
                
                columns_json.append(col_info)
                           # Populate rows JSON with forward-filled labels only if summary sheet
            rows_json = []
            if s_type == "summary_report":
                last_label_vals = {col: None for col in row_header_cols}
                
                # Get hierarchy info if available
                hierarchy = tbl.get("hierarchy", [])
                hierarchy_by_row = {h["row"]: h for h in hierarchy} if hierarchy else {}
                
                all_table_rows = title_rows + header_rows + data_rows + total_rows + check_rows
                all_table_rows.sort()
                
                for r in all_table_rows:
                    row_type = "data"
                    if r in title_rows:
                        row_type = "title"
                    elif r in header_rows:
                        row_type = "header"
                    elif r in total_rows:
                        row_type = "total"
                    elif r in check_rows:
                        row_type = "check"
                        
                    values_dict = {}
                    for c_idx, col_name in enumerate(headers, col_start):
                        val = ws_val.cell(row=r, column=c_idx).value
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            values_dict[col_name] = val.isoformat()
                        else:
                            values_dict[col_name] = val
                            
                    # Forward-fill row labels
                    row_label_vals = {}
                    if row_type == "data":
                        for col_name in row_header_cols:
                            orig_val = values_dict.get(col_name)
                            if orig_val is not None and str(orig_val).strip() != "":
                                last_label_vals[col_name] = orig_val
                            row_label_vals[col_name] = last_label_vals[col_name]
                    else:
                        first_cell = ws_val.cell(row=r, column=col_start).value
                        row_label_vals = {row_header_cols[0]: first_cell} if row_header_cols else {}
                    
                    # Add hierarchy path if available
                    hierarchy_path = ""
                    if r in hierarchy_by_row:
                        hierarchy_path = hierarchy_by_row[r].get("full_path", "")
                        
                    row_entry = {
                        "excel_row": r,
                        "row_label_values": row_label_vals,
                        "values": values_dict,
                        "row_type": row_type,
                    }
                    if hierarchy_path:
                        row_entry["hierarchy_path"] = hierarchy_path
                    rows_json.append(row_entry)
            
            # Detect input cells (manual vs formula-driven)
            input_cells = formatting_extractor.detect_input_cells(
                ws_val, ws_form, 
                data_rows[0] if data_rows else row_start,
                data_rows[-1] if data_rows else row_end,
                col_start, col_end
            )
            
            table_entry = {
                "table_name": t_name,
                "table_type": t_type,
                "section_title": section_title,
                "table_range": tbl["table_range"],
                "header_row": header_rows[-1] if header_rows else row_start,
                "data_start_row": data_rows[0] if data_rows else row_start,
                "data_end_row": data_rows[-1] if data_rows else row_end,
                "total_rows": total_rows,
                "check_rows": check_rows,
                "row_header_columns": row_header_cols,
                "column_header_rows": header_rows[:-1] if len(header_rows) > 1 else [],
                "row_count": len(data_rows),
                "column_count": len(headers),
                "input_cell_count": len(input_cells),
                "definition": t_sem.get("definition") or f"Summary table for {t_name}",
                "inter_table_relationships": t_sem.get("inter_table_relationships", []),
                "columns": columns_json,
                "rows": rows_json,
            }
            
            # Add hierarchy summary if detected
            if hierarchy:
                indent_levels = set(h.get("indent_level", 0) for h in hierarchy)
                if len(indent_levels) > 1:
                    table_entry["has_row_hierarchy"] = True
                    table_entry["hierarchy_depth"] = max(indent_levels)
            
            tables_json.append(table_entry)
            
        sheet_meta_dict = {
            "row_count": row_count,
            "column_count": col_count,
            "formula_count": formula_count,
            "non_empty_cells": non_empty_count,
            "table_count": len(sheet_tables),
            "pivot_table_count": len(pivot_tables_meta) if s_type == "summary_report" else 0,
            "print_area": sheet_meta.get("print_area"),
            "hidden_row_count": len(sheet_meta.get("hidden_rows", [])),
            "hidden_column_count": len(sheet_meta.get("hidden_columns", [])),
        }
        if s_type != "summary_report":
            sheet_meta_dict["columns"] = columns_list
            
        sheets_json.append({
            "sheet_name": s_name,
            "sheet_type": s_type,
            "sheet_range": f"A1:{get_column_letter(col_count or 1)}{row_count or 1}",
            "sheet_metadata": sheet_meta_dict,
            "filters": filters_list,
            "pivot_tables": pivot_tables_meta if s_type == "summary_report" else [],
            "tables": tables_json,
        })
        
    for sheet_json in sheets_json:
        normalize_pivot_metadata_only(sheet_json)
        
    final_json = {
        "schema_version": "6.0",
        "file_name": os.path.basename(file_name),
        "generated_at": datetime.datetime.now().isoformat(),
        "purpose": workbook_purpose or "Excel summary/report extraction with raw-data lineage and formatting intelligence",
        "process_flow": process_flow,
        "workbook_metadata": {
            "file_name": os.path.basename(file_name),
            "file_hash_md5": file_hash,
            "sheet_count": len(sheet_classifications),
            "sheet_names": list(sheet_classifications.keys()),
            "raw_data_sheet_name": raw_data_sheet_name,
            "summary_sheet_name": summary_sheet_name,
            "named_ranges": wb_meta.get("named_ranges", []),
            "external_links": wb_meta.get("external_links", []),
            "has_vba_macros": wb_meta.get("has_vba_macros", False),
            "vba_macro_streams": wb_meta.get("vba_macro_streams", []),
        },
        "sheets": sheets_json,
    }
    
    return final_json
