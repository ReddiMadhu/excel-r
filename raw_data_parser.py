"""
Raw Data Parser — Parse raw/source data sheets to extract column metadata.

Enhanced to use formatting (bold header row) as primary header detection
instead of relying solely on text content heuristics.
"""
import openpyxl
import datetime


def detect_header_row(ws):
    """
    Detect which row contains column headers.
    
    Primary: Look for a bold row with multiple text values.
    Fallback: First row with multiple unique text values (original logic).
    """
    max_inspect_rows = min((ws.max_row or 1) + 1, 20)
    max_inspect_cols = min((ws.max_column or 1) + 1, 50)
    
    # Primary: Find bold header row
    for r in range(1, max_inspect_rows):
        vals = []
        bold_count = 0
        for c in range(1, max_inspect_cols):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                vals.append(cell.value)
                if cell.font and cell.font.bold:
                    bold_count += 1
        
        non_empty = [v for v in vals if v is not None]
        if len(non_empty) >= 3:
            # If majority of cells are bold, this is likely the header
            if bold_count > 0 and bold_count / len(non_empty) > 0.5:
                str_count = sum(1 for v in non_empty if isinstance(v, str))
                if str_count / len(non_empty) >= 0.5:
                    return r
    
    # Fallback: First row with multiple text values (original logic)
    for r in range(1, max_inspect_rows):
        vals = [ws.cell(row=r, column=c).value for c in range(1, max_inspect_cols)]
        non_empty = [v for v in vals if v is not None]
        if len(non_empty) >= 3:
            str_count = sum(1 for v in non_empty if isinstance(v, str))
            if str_count / len(non_empty) >= 0.5:
                return r
                
    return 1  # Default fallback


def get_raw_columns(ws, header_row=1):
    """Return list of column headers from the header row."""
    cols = []
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is not None:
            cols.append(str(val).strip())
        else:
            cols.append(f"Column_{openpyxl.utils.get_column_letter(c)}")
    return cols


def map_excel_column_to_header(ws, header_row=1):
    """Return dictionary mapping column letters to their header name."""
    mapping = {}
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        val = ws.cell(row=header_row, column=c).value
        if val is not None:
            mapping[col_letter] = str(val).strip()
        else:
            mapping[col_letter] = f"Column_{col_letter}"
    return mapping


def infer_data_type(values):
    """Infer the general data type of a list of values."""
    if not values:
        return "string"
    types = set()
    for v in values:
        if v is None:
            continue
        if isinstance(v, bool):
            types.add("boolean")
        elif isinstance(v, int):
            types.add("integer")
        elif isinstance(v, float):
            types.add("float")
        elif isinstance(v, (datetime.datetime, datetime.date)):
            types.add("date")
        else:
            v_str = str(v).strip()
            if not v_str:
                continue
            try:
                int(v_str)
                types.add("integer")
            except ValueError:
                try:
                    float(v_str)
                    types.add("float")
                except ValueError:
                    types.add("string")
                    
    if "string" in types:
        return "string"
    if "float" in types:
        return "float"
    if "integer" in types:
        return "integer"
    if "date" in types:
        return "date"
    if "boolean" in types:
        return "boolean"
    return "string"


def get_column_samples_and_types(ws, header_row=1, max_samples=20):
    """
    For each column, fetch up to max_samples non-empty values
    and infer the column data type.
    """
    columns_metadata = {}
    col_names = get_raw_columns(ws, header_row)
    
    start_row = header_row + 1
    max_row = ws.max_row or start_row
    
    for idx, col_name in enumerate(col_names, 1):
        col_vals = []
        for r in range(start_row, min(max_row + 1, start_row + 2000)):
            val = ws.cell(row=r, column=idx).value
            if val is not None:
                if isinstance(val, (datetime.datetime, datetime.date)):
                    col_vals.append(val.isoformat())
                else:
                    col_vals.append(val)
                if len(col_vals) >= max_samples:
                    break
        
        dtype = infer_data_type(col_vals)
        columns_metadata[col_name] = {
            "excel_column": openpyxl.utils.get_column_letter(idx),
            "data_type": dtype,
            "sample_values": col_vals
        }
        
    return columns_metadata
