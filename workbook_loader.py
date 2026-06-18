"""
Workbook Loader — Load Excel workbooks and extract metadata.

Enhanced to also extract:
- Named ranges
- External links
- Print areas
- Hidden rows/columns
"""
import openpyxl
import hashlib
import os


def compute_md5(path):
    """Calculate MD5 hash of a file."""
    hash_md5 = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


def load_workbook_values(path):
    """Load workbook with data_only=True to get evaluated values."""
    return openpyxl.load_workbook(path, data_only=True, read_only=False)


def load_workbook_formulas(path):
    """Load workbook with data_only=False to get formulas."""
    return openpyxl.load_workbook(path, data_only=False, read_only=False)


def get_sheet_used_range(ws):
    """Return the cell range of the sheet, e.g., 'A1:H50'."""
    min_col = ws.min_column
    min_row = ws.min_row
    max_col = ws.max_column
    max_row = ws.max_row
    if max_col is None or max_row is None or max_col < min_col or max_row < min_row:
        return "A1:A1"
    start_letter = openpyxl.utils.get_column_letter(min_col)
    end_letter = openpyxl.utils.get_column_letter(max_col)
    return f"{start_letter}{min_row}:{end_letter}{max_row}"


def get_non_empty_cells(ws):
    """Count non-empty cells in the sheet."""
    count = 0
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                count += 1
    return count


def get_formula_count(ws_formula):
    """Count number of cells containing formulas in the sheet."""
    count = 0
    for row in ws_formula.iter_rows(values_only=True):
        for val in row:
            if val is not None and str(val).startswith('='):
                count += 1
    return count


def extract_workbook_metadata(wb, file_path=None):
    """
    Extract workbook-level metadata that aids in understanding context.
    """
    meta = {
        "named_ranges": [],
        "external_links": [],
        "has_vba_macros": False,
        "vba_macro_streams": [],
    }
    
    # Extract VBA macros
    if file_path and os.path.exists(file_path):
        try:
            from oletools.olevba import VBA_Parser
            parser = VBA_Parser(file_path)
            if parser.detect_vba_macros():
                meta["has_vba_macros"] = True
                for sub_f, stream_path, vba_filename, vba_code in parser.extract_macros():
                    meta["vba_macro_streams"].append({
                        "stream_path": str(stream_path),
                        "vba_filename": str(vba_filename),
                        "code_length": len(vba_code),
                    })
        except Exception:
            pass

    # Named ranges
    try:
        if wb.defined_names:
            for name, dn in wb.defined_names.items():
                try:
                    meta["named_ranges"].append({
                        "name": name,
                        "value": dn.value,
                    })
                except Exception:
                    pass
    except Exception:
        pass
    
    # External links (references to other workbooks)
    try:
        if hasattr(wb, '_external_links') and wb._external_links:
            for link in wb._external_links:
                try:
                    if hasattr(link, 'file_link') and link.file_link:
                        meta["external_links"].append(str(link.file_link))
                    elif hasattr(link, 'Target'):
                        meta["external_links"].append(str(link.Target))
                except Exception:
                    pass
    except Exception:
        pass
    
    return meta



def extract_sheet_metadata(ws):
    """
    Extract sheet-level metadata including print area and hidden rows/columns.
    """
    meta = {
        "print_area": None,
        "hidden_rows": [],
        "hidden_columns": [],
    }
    
    # Print area — indicates what the user considers the "final output"
    try:
        if ws.print_area:
            meta["print_area"] = ws.print_area
    except Exception:
        pass
    
    # Hidden rows
    try:
        for row_idx, rd in ws.row_dimensions.items():
            if rd.hidden:
                meta["hidden_rows"].append(row_idx)
    except Exception:
        pass
    
    # Hidden columns
    try:
        for col_letter, cd in ws.column_dimensions.items():
            if cd.hidden:
                meta["hidden_columns"].append(col_letter)
    except Exception:
        pass
    
    return meta
