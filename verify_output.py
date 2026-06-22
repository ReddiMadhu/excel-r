"""
Verification script: Cross-reference JSON output against Excel formulas.
Confirms the multi-formula merge logic is working correctly.
"""
import json
import openpyxl
from openpyxl.utils import get_column_letter

# ── 1. Load the generated JSON ──
with open("data/output/LNBAR Worksite A Reserve Details.json", "r", encoding="utf-8") as f:
    output = json.load(f)

print("=" * 70)
print("VERIFICATION REPORT")
print("=" * 70)

# ── 2. List all sheets and tables ──
for sheet in output.get("sheets", []):
    sname = sheet["sheet_name"]
    tables = sheet.get("tables", [])
    print(f"\nSheet: {sname}  ({len(tables)} tables)")
    for t in tables:
        tname = t["table_name"]
        cols = t.get("columns", [])
        formula_cols = [c for c in cols if c.get("type") == "formula_based"]
        lineage_cols = [c for c in formula_cols if c.get("formula_lineage")]
        print(f"  Table: {tname}")
        print(f"    Total columns: {len(cols)}")
        print(f"    Formula columns: {len(formula_cols)}")
        print(f"    Columns with lineage: {len(lineage_cols)}")
        for fc in formula_cols:
            fl = fc.get("formula_lineage", {})
            inputs = fl.get("direct_inputs", [])
            raw_sources = fl.get("ultimate_raw_sources", [])
            print(f"      [{fc['column_name']}] type={fl.get('computation_type','N/A')}, "
                  f"direct_inputs={len(inputs)}, raw_sources={len(raw_sources)}")
            # Show the raw source columns
            for rs in raw_sources:
                print(f"        -> {rs}")

print()
print("=" * 70)
print("EXCEL FORMULA CROSS-REFERENCE")
print("=" * 70)

# ── 3. Load the Excel workbook (formulas mode) ──
wb_form = openpyxl.load_workbook("data/input/LNBAR Worksite A Reserve Details.xlsx", data_only=False)
wb_val  = openpyxl.load_workbook("data/input/LNBAR Worksite A Reserve Details.xlsx", data_only=True)

ws_form = wb_form["Summary"]
ws_val  = wb_val["Summary"]

# Print all formulas in the Summary sheet
print("\nAll formulas in Summary sheet:")
for row in range(1, ws_form.max_row + 1):
    for col in range(1, ws_form.max_column + 1):
        cell = ws_form.cell(row=row, column=col)
        if cell.value and str(cell.value).startswith("="):
            col_letter = get_column_letter(col)
            val_cell = ws_val.cell(row=row, column=col)
            print(f"  {col_letter}{row}: {cell.value}")
            print(f"          value = {val_cell.value}")

# ── 4. Specific check: Look at the last table's formulas (Exhibit 5 equivalent) ──
summary_sheet = None
for sheet in output.get("sheets", []):
    if sheet["sheet_name"] == "Summary":
        summary_sheet = sheet
        break

if summary_sheet:
    tables = summary_sheet.get("tables", [])
    if tables:
        last_table = tables[-1]
        print(f"\n{'='*70}")
        print(f"DETAILED CHECK: Last table = '{last_table['table_name']}'")
        print(f"Range: {last_table.get('table_range', 'N/A')}")
        print(f"{'='*70}")
        
        for col in last_table.get("columns", []):
            if col.get("type") == "formula_based":
                print(f"\n  Column: {col['column_name']}")
                print(f"    Formula: {col.get('formula', 'N/A')}")
                print(f"    Pattern: {col.get('formula_pattern', 'N/A')}")
                fl = col.get("formula_lineage", {})
                if fl:
                    print(f"    Computation: {fl.get('computation_type', 'N/A')}")
                    print(f"    Direct inputs ({len(fl.get('direct_inputs', []))}):")
                    for di in fl.get("direct_inputs", []):
                        nested = " (HAS NESTED)" if di.get("nested_lineage") else ""
                        print(f"      - {di.get('table','?')}::{di.get('column','?')} "
                              f"(role={di.get('role','?')}, is_raw={di.get('is_raw','?')}){nested}")
                    print(f"    Ultimate raw sources ({len(fl.get('ultimate_raw_sources', []))}):")
                    for urs in fl.get("ultimate_raw_sources", []):
                        print(f"      - {urs}")

# ── 5. Cross-check: Verify formulas per row in the key table area ──
print(f"\n{'='*70}")
print("ROW-LEVEL FORMULA SCAN (Summary sheet, rows 38-45)")
print("=" * 70)
for row in range(35, 46):
    row_data = []
    for col in range(1, 7):
        cell_form = ws_form.cell(row=row, column=col)
        cell_val = ws_val.cell(row=row, column=col)
        cl = get_column_letter(col)
        if cell_form.value and str(cell_form.value).startswith("="):
            row_data.append(f"  {cl}{row}: {cell_form.value}")
        elif cell_val.value is not None:
            row_data.append(f"  {cl}{row}: '{cell_val.value}'")
    if row_data:
        print(f"Row {row}:")
        for rd in row_data:
            print(rd)

wb_form.close()
wb_val.close()

print(f"\n{'='*70}")
print("VERIFICATION COMPLETE")
print("=" * 70)
